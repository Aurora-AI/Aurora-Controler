"""Testes Phase A2 — Graph Builder + Cycle Detection + Topological Sort"""
import sys
from pathlib import Path
import tempfile
from openpyxl import Workbook

# Setup paths
REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))


from kernel.phase_a1.extractor import extract_structure
from kernel.phase_a1_5.normalizer import normalize_workbook
from product_a.phase_a2.graph_builder import build_dag, detect_cycles, topological_sort
from product_a.trustware.pipeline_contracts import ExecutionDAG, DAGNode, DAGEdge, CycleType


def test_simple_dag():
    """Grafo acíclico simples: A1 → A2 → A3."""
    nodes = [
        DAGNode(id='S!A1', sheet='S', coordinate='A1', data_type='n'),
        DAGNode(id='S!A2', sheet='S', coordinate='A2', formula_raw='=A1*2', data_type='e'),
        DAGNode(id='S!A3', sheet='S', coordinate='A3', formula_raw='=SUM(A1:A2)', data_type='e'),
    ]
    edges = [
        DAGEdge(source='S!A1', target='S!A2'),
        DAGEdge(source='S!A1', target='S!A3'),
        DAGEdge(source='S!A2', target='S!A3'),
    ]
    dag = ExecutionDAG(nodes=nodes, edges=edges)
    dag.cycles = detect_cycles(dag)
    dag.has_cycles = len(dag.cycles) > 0
    dag.topological_order = topological_sort(dag)
    
    assert dag.has_cycles == False
    a1_idx = dag.topological_order.index('S!A1')
    a2_idx = dag.topological_order.index('S!A2')
    a3_idx = dag.topological_order.index('S!A3')
    assert a1_idx < a2_idx < a3_idx


def test_self_reference():
    """A1 depende de A1."""
    nodes = [
        DAGNode(id='S!A1', sheet='S', coordinate='A1', formula_raw='=A1+1', data_type='e'),
    ]
    edges = [
        DAGEdge(source='S!A1', target='S!A1'),
    ]
    dag = ExecutionDAG(nodes=nodes, edges=edges)
    dag.cycles = detect_cycles(dag)
    dag.has_cycles = len(dag.cycles) > 0
    assert dag.has_cycles
    assert any(c.cycle_type == CycleType.SELF_REFERENCE for c in dag.cycles)


def test_direct_cycle():
    """A1 → B1 → A1."""
    nodes = [
        DAGNode(id='S!A1', sheet='S', coordinate='A1', formula_raw='=B1*2', data_type='e'),
        DAGNode(id='S!B1', sheet='S', coordinate='B1', formula_raw='=A1+1', data_type='e'),
    ]
    edges = [
        DAGEdge(source='S!B1', target='S!A1'),
        DAGEdge(source='S!A1', target='S!B1'),
    ]
    dag = ExecutionDAG(nodes=nodes, edges=edges)
    dag.cycles = detect_cycles(dag)
    dag.has_cycles = len(dag.cycles) > 0
    assert dag.has_cycles
    assert any(c.cycle_type == CycleType.DIRECT for c in dag.cycles)


def test_diamond_pattern_acyclic():
    """A1 -> B1 -> D1 and A1 -> C1 -> D1. Should NOT detect cycles."""
    nodes = [
        DAGNode(id='S!A1', sheet='S', coordinate='A1', data_type='n'),
        DAGNode(id='S!B1', sheet='S', coordinate='B1', formula_raw='=A1', data_type='e'),
        DAGNode(id='S!C1', sheet='S', coordinate='C1', formula_raw='=A1', data_type='e'),
        DAGNode(id='S!D1', sheet='S', coordinate='D1', formula_raw='=B1+C1', data_type='e'),
    ]
    edges = [
        DAGEdge(source='S!A1', target='S!B1'),
        DAGEdge(source='S!A1', target='S!C1'),
        DAGEdge(source='S!B1', target='S!D1'),
        DAGEdge(source='S!C1', target='S!D1'),
    ]
    dag = ExecutionDAG(nodes=nodes, edges=edges)
    dag.cycles = detect_cycles(dag)
    assert not dag.has_cycles, f"Detected cycles in acyclic diamond pattern: {dag.cycles}"


def test_integration_full_pipeline():
    """Testa integração total com workbook real gerado em memória."""
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Final'
        ws['A1'] = 100
        ws['A2'] = 200
        ws['A3'] = '=SUM(A1:A2)'
        ws['B1'] = '=A3*1.1'
        ws['B2'] = '=IF(B1>200,"HIGH","LOW")'
        path = tp / 'final.xlsx'
        wb.save(path)
        
        raw = extract_structure(path)
        norm = normalize_workbook(raw)
        dag = build_dag(norm)
        
        assert len(dag.nodes) >= 5
        assert not dag.has_cycles
        assert len(dag.topological_order) >= 5
        
        order = dag.topological_order
        a1 = order.index('Final!A1')
        a2 = order.index('Final!A2')
        a3 = order.index('Final!A3')
        b1 = order.index('Final!B1')
        b2 = order.index('Final!B2')
        assert a1 < a3 and a2 < a3
        assert a3 < b1 and a3 < b2


if __name__ == "__main__":
    try:
        test_simple_dag()
        test_self_reference()
        test_direct_cycle()
        test_diamond_pattern_acyclic()
        test_integration_full_pipeline()
        print("[TEST A2.5] Todos os testes passaram com sucesso.")
    except Exception as e:
        print(f"[TEST A2.5] FALHA: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
