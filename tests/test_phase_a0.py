"""
EXRS Phase A0 — Testes Unitários do Classifier
"""
import sys
from pathlib import Path

# Garantir que o repositório esteja no path
REPO_ROOT = Path(__file__).resolve().parents[1]


from kernel.phase_a0.classifier import classify_workbook, classify_formula
from libs.trustware.pipeline_contracts import WorkbookClass, CompileDecision
import openpyxl
import tempfile


def _make_workbook(sheets_config: dict, save_path: Path) -> Path:
    """
    sheet_config: dict com sheet_name -> list of (cell, value)
    Ex: {"Sheet1": [("A1", 10), ("A2", "=SUM(A1:A2)")]}
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    for sheet_name, cells in sheets_config.items():
        ws = wb.create_sheet(title=sheet_name)
        for cell_ref, value in cells:
            ws[cell_ref] = value
    
    wb.save(save_path)
    return save_path


def test_supported_pure():
    """Workbook apenas com valores estáticos e fórmulas padrão deve ser SUPPORTED."""
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        path = _make_workbook({
            "Sheet1": [
                ("A1", 10),
                ("A2", 20),
                ("A3", "=SUM(A1:A2)"),
                ("B1", "Revenue"),
            ]
        }, tp / "supported.xlsx")
        
        report = classify_workbook(path)
        assert report.workbook_class == WorkbookClass.SUPPORTED, f"Esperado SUPPORTED, obtido {report.workbook_class}. Reasons: {report.escalate_reasons}"
        assert report.compile_decision == CompileDecision.PROCEED


def test_restricted_volatile():
    """Workbook com função volátil deve ser RESTRICTED."""
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        path = _make_workbook({
            "Sheet1": [
                ("A1", "=NOW()"),
            ]
        }, tp / "restricted.xlsx")
        
        report = classify_workbook(path)
        assert report.workbook_class == WorkbookClass.RESTRICTED, f"Esperado RESTRICTED, obtido {report.workbook_class}"
        assert report.compile_decision == CompileDecision.PROCEED_WITH_RESTRICTIONS
        assert report.restricted_items is not None
        assert len(report.restricted_items) > 0


def test_restricted_offset():
    """OFFSET é volátil e deve ser RESTRICTED."""
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        path = _make_workbook({
            "Sheet1": [
                ("A1", "=OFFSET(B1, 0, 0)"),
            ]
        }, tp / "offset.xlsx")
        
        report = classify_workbook(path)
        assert report.workbook_class == WorkbookClass.RESTRICTED


def test_classify_formula_edge_cases():
    """Testes de borda para classify_formula."""
    # Falso positivo: 'NOTE' não deve ser confundido com 'NOW'
    assert classify_formula("=NOTE(A1)") == "SUPPORTED"
    # None é supported
    assert classify_formula(None) == "SUPPORTED"
    # String vazia é valor estático, não fórmula
    assert classify_formula("") == "SUPPORTED"
    # Fórmula complexa com função volátil aninhada
    assert classify_formula("=SUM(OFFSET(B1, 0, 0))") == "RESTRICTED"
    assert classify_formula("=IF(A1>0, NOW(), 0)") == "RESTRICTED"


def test_compile_decision_mapping():
    """Verificar mapeamento correto de WorkbookClass -> CompileDecision."""
    assert CompileDecision.PROCEED.value == "proceed"
    assert CompileDecision.PROCEED_WITH_RESTRICTIONS.value == "proceed_with_restrictions"
    assert CompileDecision.ESCALATE.value == "escalate"


def test_classify_workbook_nao_deixa_handle_zip_vazado(tmp_path):
    """Regressão (30/07/2026, planilha hostil `retail_hostile_test_v1.xlsx`):
    `classify_workbook` abre o workbook com `keep_vba=True`, o que faz o openpyxl
    manter um SEGUNDO ZipFile (`wb.vba_archive`, modo 'a' sobre um BytesIO). O
    `wb.close()` fechava só o principal; o `__del__` do vba_archive rodava depois no
    coletor de lixo, tentava `fp.seek()` no arquivo já fechado e cuspia
    `ValueError: I/O operation on closed file` como *unraisable*.

    Não derrubava o processo (é "Exception ignored"), mas sujava a saída de TODO
    comando `exrs diagnose` e virava `PytestUnraisableExceptionWarning` na suíte.
    Este teste captura o hook de unraisable e força o GC — se o handle voltar a
    vazar, falha aqui em vez de virar ruído silencioso na saída do cliente."""
    import gc

    caminho = _make_workbook({"Dados": [("A1", 10), ("A2", "=A1*2")]}, tmp_path / "wb.xlsx")

    capturadas: list[str] = []
    hook_original = sys.unraisablehook
    sys.unraisablehook = lambda arg: capturadas.append(
        f"{arg.exc_type.__name__}: {arg.exc_value}"
    )
    try:
        classify_workbook(caminho)
        gc.collect()  # força o __del__ do ZipFile a rodar AGORA, não em momento aleatório
    finally:
        sys.unraisablehook = hook_original

    assert not capturadas, (
        "handle de ZipFile vazou em classify_workbook (vba_archive não foi fechado): "
        + "; ".join(capturadas)
    )


if __name__ == "__main__":
    try:
        test_supported_pure()
        test_restricted_volatile()
        test_restricted_offset()
        test_classify_formula_edge_cases()
        test_compile_decision_mapping()
        print("[TEST A0.4] Todos os testes unitários passaram.")
    except Exception as e:
        print(f"[FAIL] Erro nos testes: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
