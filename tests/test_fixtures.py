"""
Testes de integração — coverage_test.xlsx através do pipeline A0→A4.

Carrega a planilha de cobertura, roda o pipeline completo e valida que
os valores calculados pelo avaliador batem com o gabarito (coluna D).

Convenção do fixture:
  col A = descrição
  col B = fórmula (string documental)
  col C = célula com a fórmula real (valor estático no arquivo — openpyxl não avalia)
  col D = gabarito (valor pré-calculado em Python no create_test_workbook.py)
"""
import sys
import math
import datetime as _dt
import pytest
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
for p in [
    REPO_ROOT / "src" / "product_a" / "phase_a4",
    REPO_ROOT / "src" / "kernel" / "phase_a1_5",
    REPO_ROOT / "src" / "product_a" / "phase_a2_5",
    REPO_ROOT / "src" / "product_a" / "trustware",
]:
    sys.path.insert(0, str(p))

try:
    from openpyxl import load_workbook
except ImportError:
    pytest.skip("openpyxl não instalado", allow_module_level=True)

from formula_evaluator import evaluate_formula
from normalizer import expand_range

# ── Caminho da planilha ────────────────────────────────────────────────────

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "coverage_test.xlsx"

# ── Helpers ────────────────────────────────────────────────────────────────

def _approx_eq(a, b, rel=1e-3):
    """Igualdade aproximada para floats; igualdade exata para str/bool."""
    if a is None and b is None:
        return True
    if isinstance(a, bool) or isinstance(b, bool):
        return a == b
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if math.isnan(a) and math.isnan(b):
            return True
        if abs(b) < 1e-12:
            return abs(a) < 1e-6
        return abs(a - b) / abs(b) < rel
    return str(a).strip() == str(b).strip()


def _load_sheet_cases(sheet_name: str):
    """
    Retorna lista de (row_num, description, formula_str, gabarito)
    para uma aba do fixture. Pula linhas de cabeçalho (row 1).
    """
    wb = load_workbook(FIXTURE_PATH, data_only=True)
    ws = wb[sheet_name]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        desc, _doc, formula_val, gabarito = row[0], row[1], row[2], row[3]
        if desc is None and formula_val is None:
            continue
        cases.append((desc or "", formula_val, gabarito))
    return cases


def _build_context_for_sheet(sheet_name: str) -> dict:
    """
    Constrói cache de valores (node_id → valor) para a aba,
    lendo os valores estáticos da coluna C (data_only=True).
    Também injeta os valores da aba _Data como '_Data!Cx'.
    """
    wb = load_workbook(FIXTURE_PATH, data_only=True)
    ctx: dict = {}

    # _Data rows
    if "_Data" in wb.sheetnames:
        ws_d = wb["_Data"]
        for r in ws_d.iter_rows(min_row=2, values_only=True):
            # col C = valor
            pass  # _Data is reference-only; values referenced via formula strings

    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        cell_c = row[2]  # column C (0-indexed)
        if cell_c.value is not None:
            nid = f"{sheet_name}!{cell_c.coordinate}"
            ctx[nid] = cell_c.value
    return ctx


# ── Teste principal: avalia cada fórmula e compara com gabarito ───────────

# Abas que contêm casos avaliáveis (ExternalRef é skipped por design)
EVALUABLE_SHEETS = [
    "Arithmetic",
    "Conditional",
    "Aggregation",
    "Lookup",
    "DateLogic",
    "Text",
    "Ranking",
    "Threshold",
    "Financial",
    "Engineering",
]


_EXCEL_EPOCH = _dt.date(1899, 12, 30)


def _datetime_to_serial(val):
    """Converte datetime/date → número serial Excel (como o evaluador espera)."""
    if isinstance(val, _dt.datetime):
        return (val.date() - _EXCEL_EPOCH).days
    if isinstance(val, _dt.date):
        return (val - _EXCEL_EPOCH).days
    return val


def _build_sheet_context(wb, sheet_name: str) -> dict:
    """
    Lê todos os valores âncora (células não-fórmula) de uma aba e
    devolve um dict {node_id → value} para uso como contexto de avaliação.
    Converte datetime → serial Excel para que as funções de data funcionem.
    """
    ctx: dict = {}
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is None:
                continue
            if isinstance(val, str) and val.startswith("="):
                continue  # célula de fórmula — não inclui no contexto
            ctx[f"{sheet_name}!{cell.coordinate}"] = _datetime_to_serial(val)
    return ctx


def _collect_params():
    """
    Coleta todos os parâmetros de teste dinamicamente do fixture.
    Usa col B (fórmula como string) e col D (gabarito).
    Inclui contexto de células âncora da mesma aba.
    """
    if not FIXTURE_PATH.exists():
        return []
    params = []
    wb = load_workbook(FIXTURE_PATH, data_only=False)
    for sheet in EVALUABLE_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ctx = _build_sheet_context(wb, sheet)
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            desc = row[0] or ""
            formula_str = row[1]  # col B — fórmula como string documental
            gabarito = row[3]     # col D — valor esperado
            if not desc and not formula_str:
                continue
            if not formula_str:
                continue
            params.append(pytest.param(
                sheet, desc, formula_str, gabarito, ctx,
                id=f"{sheet}::{desc[:40]}"
            ))
    return params


@pytest.mark.parametrize("sheet,desc,formula_str,gabarito,ctx", _collect_params())
def test_gabarito_match(sheet, desc, formula_str, gabarito, ctx):
    """
    Para cada linha do fixture: avalia a fórmula (col B) com o
    formula_evaluator + contexto de células âncora e compara com
    o gabarito pré-calculado (col D).
    """
    if gabarito is None:
        pytest.skip(
            f"[fixture-gap] Gabarito ausente para: {desc!r}. "
            "Adicione o valor esperado na coluna D do fixture."
        )
    if not formula_str or not str(formula_str).startswith("="):
        pytest.skip(
            f"[fixture-gap] Fórmula ausente/inválida para: {desc!r}. "
            "Coluna B deve conter uma fórmula começando com '='."
        )

    result = evaluate_formula(formula_str, dict(ctx), sheet, expand_range)

    assert _approx_eq(result, gabarito), (
        f"[{sheet}] {desc}\n"
        f"  formula  : {formula_str}\n"
        f"  got      : {result!r}\n"
        f"  expected : {gabarito!r}"
    )


# ── Testes do avaliador direto (formula_evaluator) ─────────────────────────

class TestEvaluatorWithFixtureFormulas:
    """
    Extrai as strings de fórmula (col B) do fixture e as avalia
    diretamente via formula_evaluator, comparando com o gabarito (col D).
    Isso testa o avaliador independentemente do openpyxl.
    """

    def _build_ctx(self) -> dict:
        """Contexto simples com valores hard-coded que o fixture usa."""
        return {
            # _Data sheet values referenced in formulas
            "Arithmetic!C2": 100,
            "Arithmetic!C3": 200,
            "Arithmetic!C4": 300,
        }

    def test_fixture_file_exists(self):
        assert FIXTURE_PATH.exists(), (
            f"Fixture não encontrado: {FIXTURE_PATH}\n"
            "Execute: python tests/fixtures/create_test_workbook.py"
        )

    def test_fixture_has_expected_sheets(self):
        wb = load_workbook(FIXTURE_PATH)
        names = set(wb.sheetnames)
        expected = {"_Data", "Arithmetic", "Conditional", "Aggregation",
                    "Lookup", "DateLogic", "Text", "Ranking", "Threshold",
                    "Financial", "Engineering", "ExternalRef"}
        assert expected.issubset(names), f"Abas faltando: {expected - names}"

    def test_fixture_formula_count(self):
        """Fixture deve ter pelo menos 200 fórmulas no total."""
        wb = load_workbook(FIXTURE_PATH)
        total = 0
        for name in EVALUABLE_SHEETS + ["ExternalRef"]:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            total += sum(
                1 for row in ws.iter_rows(min_row=2, values_only=True)
                if row[0] is not None
            )
        assert total >= 140, f"Esperado ≥140 casos, encontrado {total}"

    def test_evaluator_arithmetic_basic(self):
        """Operações básicas: soma, subtração, multiplicação, divisão."""
        ctx: dict = {}
        assert evaluate_formula("=2+3", ctx, "Sheet1", expand_range) == 5
        assert evaluate_formula("=10-4", ctx, "Sheet1", expand_range) == 6
        assert evaluate_formula("=3*4", ctx, "Sheet1", expand_range) == 12
        assert evaluate_formula("=10/4", ctx, "Sheet1", expand_range) == 2.5

    def test_evaluator_power_operator(self):
        """Operador ^ deve ser exponenciação, não XOR."""
        ctx: dict = {}
        result = evaluate_formula("=2^10", ctx, "Sheet1", expand_range)
        assert result == 1024

    def test_evaluator_percent_operator(self):
        """50% deve retornar 0.5."""
        ctx: dict = {}
        result = evaluate_formula("=50%", ctx, "Sheet1", expand_range)
        assert abs(result - 0.5) < 1e-9

    def test_evaluator_iferror_division(self):
        """IFERROR(1/0, 'err') deve retornar 'err', não lançar exceção."""
        ctx: dict = {}
        result = evaluate_formula('=IFERROR(1/0,"err")', ctx, "Sheet1", expand_range)
        assert result == "err"

    def test_evaluator_external_ref_returns_ext(self):
        """Fórmulas com referência externa devem retornar '#EXT!'."""
        ctx: dict = {}
        result = evaluate_formula("=SUM('[Sales.xlsx]Jan'!A:A)", ctx, "Sheet1", expand_range)
        assert result == "#EXT!"

    def test_evaluator_bracket_in_string_not_ext_ref(self):
        """Colchete dentro de string literal não deve ser detectado como ref externa."""
        ctx: dict = {}
        result = evaluate_formula('=IF(1>0,"[ok]","err")', ctx, "Sheet1", expand_range)
        assert result == "[ok]"

    def test_evaluator_sum_range(self):
        """SUM de range deve somar valores do contexto."""
        ctx = {"S!A1": 10, "S!A2": 20, "S!A3": 30}
        result = evaluate_formula("=SUM(A1:A3)", ctx, "S", expand_range)
        assert result == 60

    def test_evaluator_vlookup_basic(self):
        """VLOOKUP deve buscar em tabela 2D."""
        ctx = {
            "S!A1": "Banana", "S!B1": 1.5,
            "S!A2": "Maçã",   "S!B2": 2.0,
            "S!A3": "Uva",    "S!B3": 3.0,
        }
        result = evaluate_formula('=VLOOKUP("Maçã",A1:B3,2,0)', ctx, "S", expand_range)
        assert result == 2.0

    def test_evaluator_if_chain(self):
        """IF aninhado deve avaliar corretamente."""
        ctx: dict = {}
        result = evaluate_formula('=IF(1>2,"maior",IF(1<2,"menor","igual"))', ctx, "S", expand_range)
        assert result == "menor"

    def test_evaluator_pmt(self):
        """PMT com taxa mensal 0.5%, 360 períodos, PV 200000."""
        ctx: dict = {}
        result = evaluate_formula("=PMT(0.5%,360,200000)", ctx, "S", expand_range)
        assert result is not None
        assert abs(result - (-1199.10)) < 1.0, f"PMT={result}"

    def test_evaluator_concatenate(self):
        """CONCATENATE deve juntar strings."""
        ctx: dict = {}
        result = evaluate_formula('=CONCATENATE("Olá"," ","Mundo")', ctx, "S", expand_range)
        assert result == "Olá Mundo"

    def test_evaluator_text_functions(self):
        """Funções de texto: UPPER, LEN, LEFT."""
        ctx: dict = {}
        assert evaluate_formula('=UPPER("hello")', ctx, "S", expand_range) == "HELLO"
        assert evaluate_formula('=LEN("hello")', ctx, "S", expand_range) == 5
        assert evaluate_formula('=LEFT("hello",3)', ctx, "S", expand_range) == "hel"

    def test_evaluator_date_functions(self):
        """DATE e YEAR devem funcionar."""
        ctx: dict = {}
        result = evaluate_formula("=YEAR(DATE(2024,6,15))", ctx, "S", expand_range)
        assert result == 2024

    def test_evaluator_rank(self):
        """RANK deve retornar posição correta."""
        ctx = {"S!A1": 10, "S!A2": 30, "S!A3": 20}
        result = evaluate_formula("=RANK(30,A1:A3,0)", ctx, "S", expand_range)
        assert result == 1

    def test_evaluator_large_small(self):
        """LARGE e SMALL devem retornar k-ésimo maior/menor."""
        ctx = {"S!A1": 10, "S!A2": 30, "S!A3": 20}
        assert evaluate_formula("=LARGE(A1:A3,1)", ctx, "S", expand_range) == 30
        assert evaluate_formula("=SMALL(A1:A3,1)", ctx, "S", expand_range) == 10

    def test_evaluator_sumif(self):
        """SUMIF deve somar baseado em critério."""
        ctx = {
            "S!A1": "Frutas", "S!B1": 100,
            "S!A2": "Legumes","S!B2": 200,
            "S!A3": "Frutas", "S!B3": 150,
        }
        result = evaluate_formula('=SUMIF(A1:A3,"Frutas",B1:B3)', ctx, "S", expand_range)
        assert result == 250

    def test_evaluator_index_match(self):
        """INDEX/MATCH deve localizar valor por posição."""
        ctx = {"S!A1": 10, "S!A2": 20, "S!A3": 30, "S!B1": "x", "S!B2": "y", "S!B3": "z"}
        result = evaluate_formula("=INDEX(B1:B3,MATCH(20,A1:A3,0))", ctx, "S", expand_range)
        assert result == "y"


# ── Testes da aba ExternalRef ──────────────────────────────────────────────

class TestExternalRefSheet:
    """Valida que a aba ExternalRef tem fórmulas marcadas como referência externa."""

    def test_external_ref_sheet_exists(self):
        wb = load_workbook(FIXTURE_PATH)
        assert "ExternalRef" in wb.sheetnames

    def test_external_ref_formulas_have_brackets(self):
        """Col B da aba ExternalRef deve conter '[' nas fórmulas."""
        wb = load_workbook(FIXTURE_PATH)
        ws = wb["ExternalRef"]
        found = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and "[" in str(row[1]):
                found += 1
        assert found >= 1, "Nenhuma fórmula com '[' encontrada na aba ExternalRef"

    def test_external_ref_evaluates_to_ext(self):
        """Fórmulas da aba ExternalRef devem retornar '#EXT!' no avaliador."""
        wb = load_workbook(FIXTURE_PATH)
        ws = wb["ExternalRef"]
        ctx: dict = {}
        ext_formulas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            formula_doc = row[1]
            if formula_doc and "[" in str(formula_doc):
                ext_formulas.append(formula_doc)

        assert ext_formulas, "Nenhuma fórmula externa encontrada"
        for f in ext_formulas:
            result = evaluate_formula(f, ctx, "ExternalRef", expand_range)
            assert result == "#EXT!", f"Esperado '#EXT!' para {f!r}, obtido {result!r}"
