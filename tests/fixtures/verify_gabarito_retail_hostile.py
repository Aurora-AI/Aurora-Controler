# -*- coding: utf-8 -*-
"""
Reverificador do GABARITO_retail_hostile_test_v1.md.

Recalcula os 22 números selados direto do `.xlsx` e falha (exit 1) se algum divergir.
O gabarito e a fixture têm de morrer juntos: se alguém tocar no arquivo, este script
acusa antes que um golden master errado entre no repo.

Uso:
    python tests/fixtures/verify_gabarito_retail_hostile.py
"""
import re
import sys
from pathlib import Path

import openpyxl

FIXTURE = Path(__file__).parent / "retail_hostile_test_v1.xlsx"
SHEET = "Vendas_Oficial_v4_final"
# Linhas 7-21 e 23-37 (0-based: 6-20 e 22-36). A linha 22 (idx 21) é o subtotal
# mesclado — trap #4, nunca é transação.
DATA_ROWS = [i for i in range(6, 37) if i != 21]

# --- gabarito selado (§2 do GABARITO_retail_hostile_test_v1.md) ---
CONSOLIDADO = {"faturamento": 36553.50, "custo": 25150.00, "d40": 38053.50,
               "lucro": 11403.50, "margem_pct": 31.2, "n_transacoes": 30}
POR_VENDEDOR = {  # nome: (n, faturamento, custo)
    "Carlos Silva": (11, 15230.50, 11000.00),
    "Ana Costa": (10, 9982.00, 6150.00),
    "Roberto M.": (4, 6401.00, 4250.00),
    "Marcos Lima": (5, 4940.00, 3750.00),
}
POR_SKU = {  # sku: (n, faturamento, custo)
    "SKU-9921": (9, 11350.00, 8900.00),
    "SKU-5541": (3, 9600.00, 6300.00),
    "SKU-8832": (8, 7153.50, 4200.00),
    "SKU-3321": (4, 6000.00, 3800.00),
    "SKU-1044": (6, 2450.00, 1950.00),
}
CONCENTRACAO_CARLOS_PCT = 41.7  # arredondado, não truncado — ver nota de precisão §3
FUDGE_FACTOR = 1500.00
TOL = 0.005


def parse_brl(v):
    """Convenções de parsing do §2.D do gabarito — parte da folha de respostas.

    Traço FINAL é padding do formato Contábil (nunca sinal); parênteses são o
    negativo contábil; traço isolado é zero.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    negativo = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[R$\s()]", "", s).rstrip("-")
    if s in ("", "-"):
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    try:
        valor = float(s)
    except ValueError:
        return None
    return -valor if negativo else valor


def main() -> int:
    if not FIXTURE.exists():
        print(f"ERRO: fixture ausente: {FIXTURE}", file=sys.stderr)
        return 1

    ws = openpyxl.load_workbook(FIXTURE, read_only=True)[SHEET]
    linhas = list(ws.iter_rows(values_only=True))

    total_d = total_f = 0.0
    vendedores: dict[str, dict] = {}
    skus: dict[str, dict] = {}
    for i in DATA_ROWS:
        r = linhas[i]
        preco = parse_brl(r[3]) or 0.0
        # F11 (referência externa) não é número — fica fora do custo, nunca estimada.
        custo = r[5] if isinstance(r[5], (int, float)) else 0.0
        total_d += preco
        total_f += custo
        for acc, chave in ((vendedores, r[2]), (skus, r[1])):
            e = acc.setdefault(chave, {"n": 0, "fat": 0.0, "custo": 0.0})
            e["n"] += 1
            e["fat"] += preco
            e["custo"] += custo

    falhas: list[str] = []

    def confere(rotulo, real, esperado, tol=TOL):
        if abs(real - esperado) > tol:
            falhas.append(f"{rotulo}: real={real:,.2f} != gabarito={esperado:,.2f}")

    # §2.A consolidado
    confere("n_transacoes", len(DATA_ROWS), CONSOLIDADO["n_transacoes"])
    confere("faturamento (ΣD)", total_d, CONSOLIDADO["faturamento"])
    confere("custo (ΣF)", total_f, CONSOLIDADO["custo"])
    confere("D40 declarado", total_d + FUDGE_FACTOR, CONSOLIDADO["d40"])
    confere("lucro bruto", total_d - total_f, CONSOLIDADO["lucro"])
    confere("margem %", 100 * (total_d - total_f) / total_d, CONSOLIDADO["margem_pct"], 0.05)

    # §2.B por vendedor
    for nome, (n, fat, custo) in POR_VENDEDOR.items():
        real = vendedores.get(nome)
        if real is None:
            falhas.append(f"vendedor ausente na fixture: {nome}")
            continue
        confere(f"{nome}.n", real["n"], n)
        confere(f"{nome}.faturamento", real["fat"], fat)
        confere(f"{nome}.custo", real["custo"], custo)
    if extras := set(vendedores) - set(POR_VENDEDOR):
        falhas.append(f"vendedores não previstos no gabarito: {sorted(extras)}")

    # §2.C por SKU
    for sku, (n, fat, custo) in POR_SKU.items():
        real = skus.get(sku)
        if real is None:
            falhas.append(f"SKU ausente na fixture: {sku}")
            continue
        confere(f"{sku}.n", real["n"], n)
        confere(f"{sku}.faturamento", real["fat"], fat)
        confere(f"{sku}.custo", real["custo"], custo)
    if extras := set(skus) - set(POR_SKU):
        falhas.append(f"SKUs não previstos no gabarito: {sorted(extras)}")

    # as três tabelas têm de fechar entre si
    confere("Σ vendedores == Σ SKUs", sum(v["fat"] for v in vendedores.values()),
            sum(s["fat"] for s in skus.values()))

    # §3 veredito
    confere("concentração Carlos Silva %",
            round(100 * vendedores["Carlos Silva"]["fat"] / total_d, 1),
            CONCENTRACAO_CARLOS_PCT, 0.05)
    pior = min(skus.items(), key=lambda kv: (kv[1]["fat"] - kv[1]["custo"]) / kv[1]["fat"])
    if pior[0] != "SKU-1044":
        falhas.append(f"pior margem: real={pior[0]} != gabarito=SKU-1044")

    # traps estruturais que são fato do arquivo (não do motor)
    wb = openpyxl.load_workbook(FIXTURE)
    ws_full = wb[SHEET]
    merges = {str(m) for m in ws_full.merged_cells.ranges}
    if "A22:C22" not in merges:
        falhas.append(f"trap #4: merge A22:C22 ausente (merges reais: {sorted(merges)})")
    if ws_full["G10"].value != "=G10/0":
        falhas.append(f"trap #3: G10 != '=G10/0' (real: {ws_full['G10'].value!r})")
    if "1500" not in str(ws_full["D40"].value):
        falhas.append(f"trap #5: fudge factor ausente em D40 (real: {ws_full['D40'].value!r})")
    if "[Estoque_2025.xlsx]" not in str(ws_full["F11"].value):
        falhas.append(f"trap #2: referência externa ausente em F11 (real: {ws_full['F11'].value!r})")

    if falhas:
        print("GABARITO DIVERGIU DA FIXTURE — nada foi selado:", file=sys.stderr)
        for f in falhas:
            print(f"  ✗ {f}", file=sys.stderr)
        return 1

    print(f"OK — gabarito confere com a fixture ({len(DATA_ROWS)} transações).")
    print(f"     Faturamento R$ {total_d:,.2f} · Custo R$ {total_f:,.2f} · "
          f"Lucro R$ {total_d - total_f:,.2f} ({100*(total_d-total_f)/total_d:.1f}%)")
    print(f"     4 traps estruturais confirmadas no arquivo (#2 F11, #3 G10, #4 A22:C22, #5 D40).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
