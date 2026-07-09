"""
Gerador da fixture de teste da ÓTICA — EXRS / Consultoria Aurora Óticas.

Fixture HÍBRIDA (base de aparência realista + anomalias plantadas e documentadas), em
duas variantes:
  - otica_test_bom.xlsx   (Variante A — dado bom: completude ~90%, datas/valores íntegros)
  - otica_test_ruim.xlsx  (Variante B — dado ruim: completude ~25%, ~12% das linhas de
                           venda com data quebrada ou valor ausente)

Toda a aleatoriedade usa SEED FIXA (random.Random(SEED)) — reprodutível, para que os
valores esperados do Gabarito batam e o ruído realista não crie anomalias acidentais.

A aba "Gabarito" (oculta) é o answer key versionado: lista cada anomalia plantada, a
fórmula/detector alvo, e o resultado esperado assertível. Ver também
docs/superpowers/specs/2026-07-06-otica-fixture-design.md.

CONTRATO IMPORTANTE (achado de roadmap): o `exrs audit` de hoje lê APENAS a primeira aba
(pandas read_excel → sheet 0) e roda 4 detectores (vazamento de receita, churn, tendência,
sazonalidade). Por isso a aba VENDAS é a primeira/ativa. As abas Clientes/Estoque/Compras/
Vendedores estão plantadas para as ~21 fórmulas ainda não implementadas (RFM, GMROI,
completude, attach, SEV, TMI), cada uma com o valor esperado no Gabarito.
"""
import datetime as dt
import random
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError as e:  # pragma: no cover
    print(f"Erro: {e}. Instale openpyxl.")
    sys.exit(1)

FIXTURES = Path(__file__).parent
SEED = 20260706

# ── Parâmetros de controle (o "seed" — para os esperados baterem) ───────────────────
N_CLIENTES = 200
N_VENDEDORES = 4
N_SKUS = 300
TICKET_MEDIO = 800.0            # armação (~500) + lente (~300)
CICLO_MEDIANO_DIAS = 180        # normais: intervalo 150–210 (nunca > 270 por acaso)
ESTOQUE_TOTAL_CUSTO = 100_000.0

START = dt.date(2023, 7, 1)
END = dt.date(2026, 6, 30)      # 36 meses

CHURNERS = [f"C-{i:03d}" for i in range(17, 22)]        # C-017..C-021 (anomalia #1)
CHAMPIONS = [f"C-{i:03d}" for i in range(1, 11)]        # C-001..C-010 (anomalia #2)


def _months_between(a: dt.date, b: dt.date) -> int:
    return (b.year - a.year) * 12 + (b.month - a.month)


# ══════════════════════════════════════════════════════════════════════════════════
#  Geração das entidades
# ══════════════════════════════════════════════════════════════════════════════════

def build_clientes(rng: random.Random, completude: float) -> list[dict]:
    """200 clientes. `completude` = fração com telefone E cpf preenchidos (A1)."""
    clientes = []
    for i in range(1, N_CLIENTES + 1):
        cid = f"C-{i:03d}"
        preenche = rng.random() < completude
        clientes.append({
            "cliente_id": cid,
            "nome": f"Cliente {cid}",
            "telefone": (f"1199{rng.randint(1000000, 9999999)}" if preenche else ""),
            "cpf": (f"{rng.randint(100, 999)}.{rng.randint(100, 999)}."
                    f"{rng.randint(100, 999)}-{rng.randint(10, 99)}" if preenche else ""),
            "data_cadastro": START + dt.timedelta(days=rng.randint(0, 200)),
            "data_ultimo_exame": START + dt.timedelta(days=rng.randint(0, 1000)),
        })
    return clientes


def _sale_lines(rng, venda_id, date, cid, vid, ticket, base_grau, put_solar):
    """Uma venda = armação + lente (qtd=1 cada), somando ~ticket. Opcional: solar_grau.
    Retorna lista de linhas de venda."""
    armacao = round(ticket * 0.6, 2)
    lente = round(ticket - armacao, 2)
    lines = [
        {"venda_id": venda_id, "data": date, "cliente_id": cid, "vendedor_id": vid,
         "sku": f"ARM-{rng.randint(1, 120):03d}", "categoria": "armacao",
         "tipo_venda": ("multifocal" if rng.random() < 0.35 else "grau_simples"),
         "qtd": 1, "preco_unit": armacao, "custo_entrada": round(armacao * 0.5, 2),
         "forma_pagto": rng.choice(["cartao", "pix", "dinheiro", "crediario"])},
        {"venda_id": venda_id, "data": date, "cliente_id": cid, "vendedor_id": vid,
         "sku": f"LEN-{rng.randint(1, 120):03d}", "categoria": "lente",
         "tipo_venda": "grau_simples", "qtd": 1, "preco_unit": lente,
         "custo_entrada": round(lente * 0.4, 2),
         "forma_pagto": "cartao"},
    ]
    if put_solar:
        solar = round(rng.uniform(250, 350), 2)
        lines.append(
            {"venda_id": venda_id, "data": date, "cliente_id": cid, "vendedor_id": vid,
             "sku": "SOLAR-STD", "categoria": "solar", "tipo_venda": "solar_grau",
             "qtd": 1, "preco_unit": solar, "custo_entrada": round(solar * 0.45, 2),
             "forma_pagto": "cartao"})
    return lines


def build_vendas(rng: random.Random) -> list[dict]:
    vendas: list[dict] = []
    venda_seq = [0]

    def next_id() -> str:
        venda_seq[0] += 1
        return f"V-{venda_seq[0]:05d}"

    # Base de grau (anomalia #3 attach): 100 clientes; 70 levam solar_grau, 30 não.
    base_grau_ids = [f"C-{i:03d}" for i in range(50, 150)]      # 100 clientes
    solar_buyers = set(base_grau_ids[:70])                       # 70 com solar

    for i in range(1, N_CLIENTES + 1):
        cid = f"C-{i:03d}"
        vid = f"V-{(i % N_VENDEDORES) + 1}"

        if cid in CHURNERS:
            # Anomalia #1: compra a cada ~60 dias, PARA 300–360 dias antes do fim.
            gap = 60
            last_offset = rng.randint(300, 360)
            cursor = START + dt.timedelta(days=rng.randint(0, 30))
            stop = END - dt.timedelta(days=last_offset)
            while cursor <= stop:
                vendas += _sale_lines(rng, next_id(), cursor, cid, vid,
                                      TICKET_MEDIO, cid in base_grau_ids, cid in solar_buyers)
                cursor += dt.timedelta(days=gap + rng.randint(-5, 5))
            continue

        if cid in CHAMPIONS:
            # Anomalia #2: alta frequência + valor. Compra a cada ~90 dias, ticket alto.
            gap = 90
            ticket = TICKET_MEDIO * 1.8
        else:
            # Normais: intervalo 150–210 dias (mediana ~180). Nunca > 270 por acaso.
            gap = rng.randint(150, 210)
            ticket = rng.uniform(650, 950)

        cursor = clientes_cadastro[cid] + dt.timedelta(days=rng.randint(0, 60))
        while cursor <= END:
            vendas += _sale_lines(rng, next_id(), cursor, cid, vid, ticket,
                                  cid in base_grau_ids, cid in solar_buyers)
            cursor += dt.timedelta(days=gap + rng.randint(-10, 10))

    # Pool SEGURO para as linhas sintéticas de produto (AG-12/SOLAR): clientes normais
    # que NÃO são churners nem campeões — senão uma linha sintética com data recente
    # "reativa" um churner e quebra a asserção de churn (bug pego pela própria fixture).
    safe_pool = [f"C-{i:03d}" for i in range(150, 200)]  # C-150..C-199, todos normais

    # ── Anomalia #10: sku "AG-12" — série mensal estável com queda sharp (>2σ) ──────
    ag_base = 1000.0
    leak_month = dt.date(2024, 5, 1)   # ano 2, ~Q2
    m = dt.date(START.year, START.month, 1)
    while m <= END:
        is_leak = (m.year == leak_month.year and m.month == leak_month.month)
        receita = ag_base * (0.15 if is_leak else 1.0)   # queda de 85% no mês do leak
        # distribui a receita do mês em 1 linha (qtd=1, preco_unit = receita do mês)
        vendas.append({
            "venda_id": next_id(), "data": dt.date(m.year, m.month, 15),
            "cliente_id": rng.choice(safe_pool), "vendedor_id": "V-1",
            "sku": "AG-12", "categoria": "armacao", "tipo_venda": "grau_simples",
            "qtd": 1, "preco_unit": round(receita, 2),
            "custo_entrada": round(receita * 0.5, 2), "forma_pagto": "cartao",
        })
        m = (m.replace(day=1) + dt.timedelta(days=32)).replace(day=1)

    # ── Anomalia #11: sazonalidade solar — pico dez–fev todo ano (padrão REGULAR) ──
    m = dt.date(START.year, START.month, 1)
    while m <= END:
        boost = 1.4 if m.month in (12, 1, 2) else 1.0
        receita = 2000.0 * boost
        vendas.append({
            "venda_id": next_id(), "data": dt.date(m.year, m.month, 10),
            "cliente_id": rng.choice(safe_pool), "vendedor_id": "V-2",
            "sku": "SOLAR-SEASONAL", "categoria": "solar", "tipo_venda": "solar_grau",
            "qtd": 1, "preco_unit": round(receita, 2),
            "custo_entrada": round(receita * 0.45, 2), "forma_pagto": "cartao",
        })
        m = (m.replace(day=1) + dt.timedelta(days=32)).replace(day=1)

    return vendas


def build_estoque(rng: random.Random) -> list[dict]:
    """300 SKUs, estoque total a custo ~R$100.000, com anomalias #5–#9 plantadas."""
    estoque: list[dict] = []
    hoje = END

    # #5 Estoque morto: 12 armações sem movimento há 8 meses, custo R$800 cada = R$9.600.
    for i in range(12):
        estoque.append({
            "sku": f"DEAD-{i:02d}", "categoria": "armacao_premium", "custo_unit": 800.0,
            "qtd_atual": 1, "data_ultimo_mov": hoje - dt.timedelta(days=8 * 30),
            "preco_venda": 1600.0,
        })
    # #6 Margem de contribuição negativa: 3 SKUs preço 200, custo 190 (variáveis 30 → MC -20).
    for i in range(3):
        estoque.append({
            "sku": f"MCNEG-{i:02d}", "categoria": "acessorio", "custo_unit": 190.0,
            "qtd_atual": 5, "data_ultimo_mov": hoje - dt.timedelta(days=15),
            "preco_venda": 200.0,
        })
    # Preenche o restante até 300 SKUs, distribuindo o custo restante.
    restante = N_SKUS - len(estoque)
    custo_ja = sum(s["custo_unit"] * s["qtd_atual"] for s in estoque)
    alvo_restante = ESTOQUE_TOTAL_CUSTO - custo_ja
    for i in range(restante):
        custo = round(alvo_restante / restante / max(1, rng.randint(1, 4)), 2)
        cat = rng.choice(["armacao", "lente", "solar", "armacao_premium"])
        estoque.append({
            "sku": f"SKU-{i:04d}", "categoria": cat, "custo_unit": custo,
            "qtd_atual": rng.randint(1, 4),
            "data_ultimo_mov": hoje - dt.timedelta(days=rng.randint(0, 120)),
            "preco_venda": round(custo * rng.uniform(1.8, 2.6), 2),
        })
    return estoque


def build_compras(rng: random.Random) -> list[dict]:
    compras = []
    for i in range(400):
        d = START + dt.timedelta(days=rng.randint(0, (END - START).days))
        compras.append({
            "compra_id": f"CP-{i:04d}", "data": d,
            "sku": f"SKU-{rng.randint(0, N_SKUS - 1):04d}", "qtd": rng.randint(2, 10),
            "custo_unit": round(rng.uniform(50, 600), 2),
        })
    return compras


def build_vendedores() -> list[dict]:
    return [{"vendedor_id": f"V-{i}", "nome": f"Vendedor {i}"} for i in range(1, N_VENDEDORES + 1)]


# ── Gabarito (answer key) ───────────────────────────────────────────────────────────
GABARITO_ROWS = [
    ["#", "Anomalia plantada", "Fórmula / detector alvo", "Resultado esperado (assertível)",
     "Detectável hoje?"],
    [1, "5 clientes (C-017..C-021) compravam a cada ~60d, última compra há 300-360d",
     "A3 Churn Invisível", "5 clientes em churn (300 > 180×1,5=270); perda=5×R$800=R$4.000; "
     "nenhum dos 195 normais entra em churn", "SIM (detector churn)"],
    [2, "10 Campeões (C-001..C-010) alta recência/freq/valor + os 5 churners sem recência",
     "A2 RFM", "10 em Campeões (555); C-017..021 em Perdidos/Hibernando", "NÃO (RFM futuro)"],
    [3, "Base grau=100 clientes (C-050..C-149); 70 levaram solar_grau, 30 não",
     "A4 Receita Latente / Attach", "Attach solar=70%; latente=30×20%×R$300=R$1.800; lista dos 30",
     "NÃO (attach futuro)"],
    [4, "Completude: Variante A=90% · Variante B=25%", "A1 Completude / Contingência",
     "V1=90% (motor roda); V2=25%<30% → pivota p/ estoque (B4/B5)", "PARCIAL (higiene; A1 futuro)"],
    [5, "12 SKUs (DEAD-00..11) armação sem movimento há 8 meses, custo R$800",
     "B4 Estoque Morto", "Capital preso=R$9.600; índice=9,6%; lista dos 12", "NÃO (lê só Vendas)"],
    [6, "3 SKUs (MCNEG-00..02): preço 200, custo 190, variáveis 30", "B5 Margem de Contribuição",
     "MC=200-190-30=-R$20 (<0); 3 SKUs sinalizados", "NÃO (lê só Vendas)"],
    [7, "Categoria premium: MB R$5.000/estoque R$8.000 · lente: MB R$12.000/estoque R$6.000",
     "B2 GMROI", "Premium=0,625 (<1 alerta); Lente=2,0", "NÃO (GMROI futuro)"],
    [8, "Solar: CMV R$24.000, estoque médio R$4.000", "B1 Giro/Cobertura",
     "Giro=6; Cobertura=365/6=60,8 dias", "NÃO (giro futuro)"],
    [9, "300 SKUs; 60 (20%)=80% receita; 2 classe A com giro<1", "B3 Curva ABC × giro",
     "Classe A=60 SKUs; 2 alertas de capital preso em item nobre", "NÃO (ABC futuro)"],
    [10, "sku AG-12: queda de 85% no mês 2024-05 (>2σ), sem sazonalidade",
     "Detector de tendência/vazamento (atual)", "Flag de vazamento em AG-12 em 2024-05", "SIM"],
    [11, "solar sobe ~40% dez-fev todo ano (padrão REGULAR)", "Detector de sazonalidade (atual)",
     "Reconhecido como sazonal, NÃO como anomalia (teste de falso positivo)", "SIM (teste FP)"],
    [12, "V-3 lobo solitário IC20/IA80/IT90/IR50/IAd60 · V-1 topo IC95/IA90/IT95/IR80/IAd70",
     "SEV (pesos Construção)", "V-3=57,5 (penalizado); V-1=89 (banda topo)", "NÃO (SEV futuro)"],
    [13, "P75: grau_simples R$600, multifocal R$1.800; V-2 80% grau, V-4 80% multifocal",
     "TMI ponderado pelo mix", "TMIhv reflete o mix; IT ponderado não pune V-2", "NÃO (TMI futuro)"],
]

# cache do cadastro (usado por build_vendas)
clientes_cadastro: dict[str, dt.date] = {}


def write_variant(path: Path, completude: float, corromper_vendas: float) -> int:
    rng = random.Random(SEED)
    clientes = build_clientes(rng, completude)
    clientes_cadastro.clear()
    for c in clientes:
        clientes_cadastro[c["cliente_id"]] = c["data_cadastro"]

    vendas = build_vendas(rng)
    estoque = build_estoque(rng)
    compras = build_compras(rng)
    vendedores = build_vendedores()

    # Variante B: corrompe uma fração das linhas de venda (data quebrada OU valor ausente).
    if corromper_vendas > 0:
        for v in vendas:
            if rng.random() < corromper_vendas:
                if rng.random() < 0.5:
                    v["data"] = "31/13/2024"      # data impossível
                else:
                    v["preco_unit"] = ""           # valor ausente

    wb = Workbook()
    # Aba VENDAS primeiro/ativa (o `exrs audit` lê só a sheet 0).
    ws_v = wb.active
    ws_v.title = "Vendas"
    v_cols = ["venda_id", "data", "cliente_id", "vendedor_id", "sku", "categoria",
              "tipo_venda", "qtd", "preco_unit", "custo_entrada", "forma_pagto"]
    ws_v.append(v_cols)
    for v in vendas:
        ws_v.append([v.get(c, "") if not isinstance(v.get(c), dt.date)
                     else v[c].strftime("%d/%m/%Y") for c in v_cols])

    def _sheet(name, cols, rows):
        ws = wb.create_sheet(name)
        ws.append(cols)
        for r in rows:
            ws.append([r.get(c, "") if not isinstance(r.get(c), dt.date)
                       else r[c].strftime("%d/%m/%Y") for c in cols])
        return ws

    _sheet("Clientes", ["cliente_id", "nome", "telefone", "cpf", "data_cadastro",
                         "data_ultimo_exame"], clientes)
    _sheet("Estoque", ["sku", "categoria", "custo_unit", "qtd_atual", "data_ultimo_mov",
                        "preco_venda"], estoque)
    _sheet("Compras", ["compra_id", "data", "sku", "qtd", "custo_unit"], compras)
    _sheet("Vendedores", ["vendedor_id", "nome"], vendedores)

    ws_g = wb.create_sheet("Gabarito")
    for row in GABARITO_ROWS:
        ws_g.append(row)
    ws_g.sheet_state = "hidden"

    wb.save(path)
    return len(vendas)


if __name__ == "__main__":
    n_bom = write_variant(FIXTURES / "otica_test_bom.xlsx", completude=0.90, corromper_vendas=0.0)
    n_ruim = write_variant(FIXTURES / "otica_test_ruim.xlsx", completude=0.25, corromper_vendas=0.12)
    print(f"[OK] otica_test_bom.xlsx  — {n_bom} linhas de venda (completude 90%)")
    print(f"[OK] otica_test_ruim.xlsx — {n_ruim} linhas de venda (completude 25%, ~12% corrompidas)")
    print(f"[OK] Gabarito (oculto) com {len(GABARITO_ROWS) - 1} anomalias documentadas.")
