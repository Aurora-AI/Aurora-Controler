"""
Gerador da planilha de teste de cobertura — EXRS
Cria coverage_test.xlsx com exemplos de todas as categorias de fórmula.

Estrutura:
  - Aba _Data      : valores base usados pelas fórmulas
  - Aba Arithmetic : operadores aritméticos e matemáticos
  - Aba Conditional: IF, IFS, IFERROR, AND, OR, SWITCH
  - Aba Aggregation: SUM, AVERAGE, COUNT, SUMIF, COUNTIF, SUMPRODUCT
  - Aba Lookup     : VLOOKUP, XLOOKUP, INDEX/MATCH, CHOOSE
  - Aba DateLogic  : DATE, DATEDIF, NETWORKDAYS, WORKDAY, EDATE
  - Aba Text       : CONCATENATE, LEFT, RIGHT, MID, UPPER, TRIM, TEXT
  - Aba Ranking    : RANK, LARGE, SMALL, PERCENTILE
  - Aba Threshold  : MIN/MAX condicionais, MAXIFS, MINIFS
  - Aba Financial  : PMT, FV, PV, NPV, IRR, EFFECT
  - Aba Engineering: BIN2DEC, HEX2DEC, CONVERT, BITAND

Cada aba tem:
  - Coluna A: descrição do caso
  - Coluna B: fórmula (como string — para documentação)
  - Coluna C: célula com a fórmula real
  - Coluna D: valor esperado (gabarito estático para testes)
"""
import sys
import math
import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Erro: {e}. Instale openpyxl: pip install openpyxl")
    sys.exit(1)

OUTPUT = Path(__file__).parent / "coverage_test.xlsx"

# ── Estilos ────────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT  = Font(bold=True, color="F8FAFC", size=10)
_LABEL_FILL   = PatternFill("solid", fgColor="F1F5F9")
_LABEL_FONT   = Font(bold=False, color="334155", size=10)
_FORMULA_FONT = Font(color="1D4ED8", size=10, italic=True)
_EXPECT_FILL  = PatternFill("solid", fgColor="F0FDF4")
_EXPECT_FONT  = Font(color="15803D", size=10, bold=True)
_MONO_FONT    = Font(name="Courier New", size=9, color="4B5563")
_THIN          = Side(style="thin", color="CBD5E1")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header(ws, title: str, color: str = "1E293B"):
    fill = PatternFill("solid", fgColor=color)
    for col in range(1, 5):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = _HEADER_FONT
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).value = "Descrição"
    ws.cell(row=1, column=2).value = "Fórmula (doc)"
    ws.cell(row=1, column=3).value = "Resultado"
    ws.cell(row=1, column=4).value = "Gabarito"
    ws.row_dimensions[1].height = 22
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    # Título da aba no topo
    ws.cell(row=1, column=1).value = f"[{title}] Descrição"


def _row(ws, row: int, desc: str, formula_str: str, formula_cell: str, expected):
    """Insere uma linha de teste."""
    cells = [
        (1, desc,         _LABEL_FONT,   _LABEL_FILL),
        (2, formula_str,  _MONO_FONT,    PatternFill()),
        (3, formula_cell, _FORMULA_FONT, PatternFill()),
        (4, expected,     _EXPECT_FONT,  _EXPECT_FILL),
    ]
    for col, val, font, fill in cells:
        c = ws.cell(row=row, column=col, value=val)
        c.font = font
        if fill.fill_type:
            c.fill = fill
        c.border = _BORDER
        c.alignment = Alignment(vertical="center", wrap_text=(col == 1))
    ws.row_dimensions[row].height = 18


# ── Aba _Data ─────────────────────────────────────────────────────────────

def _build_data(wb: Workbook):
    ws = wb.create_sheet("_Data")
    ws.sheet_state = "visible"
    data = [
        ("A", "Vendas Jan",    1000),
        ("A", "Vendas Fev",    2000),
        ("A", "Vendas Mar",    1500),
        ("A", "Vendas Abr",    3000),
        ("A", "Vendas Mai",     500),
        ("B", "Categoria",  "Frutas"),
        ("B", "Categoria",  "Legumes"),
        ("B", "Categoria",  "Frutas"),
        ("B", "Categoria",  "Cereais"),
        ("B", "Categoria",  "Frutas"),
        ("C", "Data Início", datetime.date(2024, 1, 1)),
        ("C", "Data Fim",    datetime.date(2024, 6, 15)),
        ("D", "Taxa mensal", 0.005),
        ("D", "Períodos",    360),
        ("D", "PV",          200000),
    ]
    ws["A1"] = "Coluna"
    ws["B1"] = "Rótulo"
    ws["C1"] = "Valor"
    for i, (col, label, val) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=col)
        ws.cell(row=i, column=2, value=label)
        ws.cell(row=i, column=3, value=val)
        if isinstance(val, datetime.date):
            ws.cell(row=i, column=3).number_format = "DD/MM/YYYY"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20


# ── Aba Arithmetic ─────────────────────────────────────────────────────────

def _build_arithmetic(wb: Workbook):
    ws = wb.create_sheet("Arithmetic")
    _header(ws, "Arithmetic", "1E40AF")
    cases = [
        # (descrição, fórmula_str, fórmula_excel, gabarito)
        ("Soma simples",              "=10+25",           "=10+25",           35),
        ("Subtração",                 "=100-37",          "=100-37",          63),
        ("Multiplicação",             "=7*8",             "=7*8",             56),
        ("Divisão",                   "=100/4",           "=100/4",           25.0),
        ("Potenciação (^)",           "=2^10",            "=2^10",            1024),
        ("Percentagem (50%)",         "=50%",             "=50%",             0.5),
        ("Módulo (MOD)",              "=MOD(17,5)",       "=MOD(17,5)",       2),
        ("Arredondamento (ROUND)",    "=ROUND(3.14159,2)","=ROUND(3.14159,2)",3.14),
        ("Arredond. p/cima (ROUNDUP)","=ROUNDUP(2.1,0)", "=ROUNDUP(2.1,0)",  3),
        ("Arredond. p/baixo",         "=ROUNDDOWN(2.9,0)","=ROUNDDOWN(2.9,0)",2),
        ("Valor absoluto (ABS)",      "=ABS(-42)",        "=ABS(-42)",        42),
        ("Raiz quadrada (SQRT)",      "=SQRT(144)",       "=SQRT(144)",       12.0),
        ("Logaritmo natural (LN)",    "=LN(EXP(1))",      "=LN(EXP(1))",      1.0),
        ("Log base 10 (LOG10)",       "=LOG10(1000)",     "=LOG10(1000)",     3.0),
        ("Seno (SIN)",                "=SIN(PI()/2)",     "=SIN(PI()/2)",     1.0),
        ("Fatorial (FACT)",           "=FACT(5)",         "=FACT(5)",         120),
        ("Combinações (COMBIN)",      "=COMBIN(5,2)",     "=COMBIN(5,2)",     10),
        ("INT trunca p/baixo",        "=INT(-2.1)",       "=INT(-2.1)",       -3),
        ("Concatenação (&)",          "=\"Hello\"&\" \"&\"World\"","=\"Hello\"&\" \"&\"World\"","Hello World"),
        ("Comparação igual (=)",      "=10=10",           "=10=10",           True),
        ("Comparação diferente (<>)", "=10<>20",          "=10<>20",          True),
        ("Divisão por zero → IFERROR","=IFERROR(1/0,\"#DIV0\")","=IFERROR(1/0,\"#DIV0\")","#DIV0"),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba Conditional ─────────────────────────────────────────────────────────

def _build_conditional(wb: Workbook):
    # Dados locais na aba
    ws = wb.create_sheet("Conditional")
    _header(ws, "Conditional", "065F46")
    # Inserir dados auxiliares nas colunas F:G (invisíveis ao leitor)
    ws["F1"] = 75   # nota
    ws["G1"] = "Aprovado"
    cases = [
        ("IF simples (verdadeiro)",   "=IF(F1>=60,\"Aprovado\",\"Reprovado\")",
                                      "=IF(F1>=60,\"Aprovado\",\"Reprovado\")", "Aprovado"),
        ("IF simples (falso)",        "=IF(F1>=90,\"Ótimo\",\"Regular\")",
                                      "=IF(F1>=90,\"Ótimo\",\"Regular\")",      "Regular"),
        ("IF aninhado",               "=IF(F1>=90,\"A\",IF(F1>=70,\"B\",\"C\"))",
                                      "=IF(F1>=90,\"A\",IF(F1>=70,\"B\",\"C\"))", "B"),
        ("AND — ambos verdadeiros",   "=AND(F1>50,F1<100)",
                                      "=AND(F1>50,F1<100)",                     True),
        ("OR — um verdadeiro",        "=OR(F1>90,F1>60)",
                                      "=OR(F1>90,F1>60)",                       True),
        ("NOT",                       "=NOT(F1=0)",
                                      "=NOT(F1=0)",                             True),
        ("XOR",                       "=XOR(F1>90,F1>60)",
                                      "=XOR(F1>90,F1>60)",                      True),
        ("IFERROR captura erro",      "=IFERROR(1/0,\"Erro\")",
                                      "=IFERROR(1/0,\"Erro\")",                 "Erro"),
        ("IFERROR sem erro",          "=IFERROR(10/2,\"Erro\")",
                                      "=IFERROR(10/2,\"Erro\")",                5.0),
        ("IFNA captura #N/A",         "=IFNA(MATCH(99,{1,2,3},0),\"NF\")",
                                      "=IFNA(MATCH(99,{1,2,3},0),\"NF\")",      "NF"),
        ("IFS múltiplas condições",   "=IFS(F1>=90,\"A\",F1>=70,\"B\",F1>=60,\"C\",TRUE,\"F\")",
                                      "=IFS(F1>=90,\"A\",F1>=70,\"B\",F1>=60,\"C\",TRUE,\"F\")", "B"),
        ("SWITCH exato",              "=SWITCH(2,1,\"um\",2,\"dois\",3,\"três\")",
                                      "=SWITCH(2,1,\"um\",2,\"dois\",3,\"três\")", "dois"),
        ("SWITCH default",            "=SWITCH(9,1,\"um\",2,\"dois\",\"outro\")",
                                      "=SWITCH(9,1,\"um\",2,\"dois\",\"outro\")", "outro"),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba Aggregation ─────────────────────────────────────────────────────────

def _build_aggregation(wb: Workbook):
    ws = wb.create_sheet("Aggregation")
    _header(ws, "Aggregation", "7C2D12")
    # Dados locais F1:F5 = [10,20,30,40,5], G1:G5 categorias
    for i, v in enumerate([10, 20, 30, 40, 5], start=1):
        ws.cell(row=i, column=6, value=v)
    for i, v in enumerate(["A","B","A","B","A"], start=1):
        ws.cell(row=i, column=7, value=v)

    cases = [
        ("SUM de range",              "=SUM(F1:F5)",          "=SUM(F1:F5)",         105),
        ("AVERAGE",                   "=AVERAGE(F1:F5)",      "=AVERAGE(F1:F5)",      21.0),
        ("COUNT (numéricos)",         "=COUNT(F1:F5)",        "=COUNT(F1:F5)",        5),
        ("COUNTA (não-vazios)",        "=COUNTA(G1:G5)",       "=COUNTA(G1:G5)",       5),
        ("MIN",                       "=MIN(F1:F5)",          "=MIN(F1:F5)",          5),
        ("MAX",                       "=MAX(F1:F5)",          "=MAX(F1:F5)",          40),
        ("COUNTIF maior que 15",      "=COUNTIF(F1:F5,\">15\")","=COUNTIF(F1:F5,\">15\")", 3),
        ("COUNTIF igual a texto",     "=COUNTIF(G1:G5,\"A\")", "=COUNTIF(G1:G5,\"A\")", 3),
        ("SUMIF categoria A",         "=SUMIF(G1:G5,\"A\",F1:F5)","=SUMIF(G1:G5,\"A\",F1:F5)", 45),
        ("AVERAGEIF categoria B",     "=AVERAGEIF(G1:G5,\"B\",F1:F5)","=AVERAGEIF(G1:G5,\"B\",F1:F5)", 30.0),
        ("COUNTIFS dois critérios",   "=COUNTIFS(F1:F5,\">10\",G1:G5,\"A\")","=COUNTIFS(F1:F5,\">10\",G1:G5,\"A\")", 1),
        ("SUMPRODUCT dois ranges",    "=SUMPRODUCT(F1:F3,F3:F5)","=SUMPRODUCT(F1:F3,F3:F5)", 10*30+20*40+30*5),
        ("SUBTOTAL(9) = SUM",         "=SUBTOTAL(9,F1:F5)",   "=SUBTOTAL(9,F1:F5)",  105),
        ("LARGE (2° maior)",          "=LARGE(F1:F5,2)",      "=LARGE(F1:F5,2)",      30),
        ("SMALL (2° menor)",          "=SMALL(F1:F5,2)",      "=SMALL(F1:F5,2)",      10),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba Lookup ─────────────────────────────────────────────────────────────

def _build_lookup(wb: Workbook):
    ws = wb.create_sheet("Lookup")
    _header(ws, "Lookup", "4C1D95")
    # Tabela de lookup em F:H
    headers = ["ID", "Nome",     "Valor"]
    data_lu = [(1, "Alpha", 100), (2, "Beta", 200), (3, "Gamma", 300), (4, "Delta", 400)]
    for j, h in enumerate(headers, start=6):
        ws.cell(row=1, column=j, value=h)
    for i, (id_, nm, vl) in enumerate(data_lu, start=2):
        ws.cell(row=i, column=6, value=id_)
        ws.cell(row=i, column=7, value=nm)
        ws.cell(row=i, column=8, value=vl)

    cases = [
        ("VLOOKUP exato (col 2)",     "=VLOOKUP(2,F2:H5,2,FALSE)", "=VLOOKUP(2,F2:H5,2,FALSE)", "Beta"),
        ("VLOOKUP exato (col 3)",     "=VLOOKUP(3,F2:H5,3,FALSE)", "=VLOOKUP(3,F2:H5,3,FALSE)", 300),
        ("XLOOKUP encontrado",        "=XLOOKUP(4,F2:F5,H2:H5)", "=XLOOKUP(4,F2:F5,H2:H5)",   400),
        ("XLOOKUP não encontrado",    "=XLOOKUP(99,F2:F5,H2:H5,\"NF\")", "=XLOOKUP(99,F2:F5,H2:H5,\"NF\")", "NF"),
        ("MATCH exato",               "=MATCH(\"Gamma\",G2:G5,0)","=MATCH(\"Gamma\",G2:G5,0)", 3),
        ("MATCH não encontrado→#N/A", "=IFERROR(MATCH(99,F2:F5,0),\"NF\")","=IFERROR(MATCH(99,F2:F5,0),\"NF\")", "NF"),
        ("INDEX 1D",                  "=INDEX(H2:H5,2)",          "=INDEX(H2:H5,2)",            200),
        ("INDEX 2D",                  "=INDEX(F2:H5,3,2)",        "=INDEX(F2:H5,3,2)",          "Gamma"),
        ("INDEX+MATCH combinados",    "=INDEX(H2:H5,MATCH(\"Beta\",G2:G5,0))","=INDEX(H2:H5,MATCH(\"Beta\",G2:G5,0))", 200),
        ("CHOOSE índice 2",           "=CHOOSE(2,\"A\",\"B\",\"C\")", "=CHOOSE(2,\"A\",\"B\",\"C\")", "B"),
        ("ROWS de range",             "=ROWS(F2:H5)",             "=ROWS(F2:H5)",               4),
        ("COLUMNS de range",          "=COLUMNS(F2:H5)",          "=COLUMNS(F2:H5)",            3),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba DateLogic ──────────────────────────────────────────────────────────

def _build_date(wb: Workbook):
    ws = wb.create_sheet("DateLogic")
    _header(ws, "DateLogic", "0C4A6E")
    # Datas auxiliares
    d_start = datetime.date(2024, 1, 1)
    d_end   = datetime.date(2024, 6, 15)
    serial_start = (d_start - datetime.date(1899, 12, 30)).days  # 45292
    serial_end   = (d_end   - datetime.date(1899, 12, 30)).days  # 45458
    ws["F1"] = serial_start
    ws["F2"] = serial_end
    ws["F1"].number_format = "DD/MM/YYYY"
    ws["F2"].number_format = "DD/MM/YYYY"

    cases = [
        ("DATE(2024,3,15)",           "=DATE(2024,3,15)",     "=DATE(2024,3,15)",     (datetime.date(2024,3,15)-datetime.date(1899,12,30)).days),
        ("YEAR de data",              "=YEAR(F1)",            "=YEAR(F1)",             2024),
        ("MONTH de data",             "=MONTH(F1)",           "=MONTH(F1)",            1),
        ("DAY de data",               "=DAY(F1)",             "=DAY(F1)",              1),
        ("DATEDIF em dias (D)",       "=DATEDIF(F1,F2,\"D\")", "=DATEDIF(F1,F2,\"D\")", (d_end-d_start).days),
        ("DATEDIF em meses (M)",      "=DATEDIF(F1,F2,\"M\")", "=DATEDIF(F1,F2,\"M\")", 5),
        ("DATEDIF em anos (Y)",       "=DATEDIF(F1,F2,\"Y\")", "=DATEDIF(F1,F2,\"Y\")", 0),
        ("DAYS360 (US method)",       "=DAYS360(F1,F2,FALSE)","=DAYS360(F1,F2,FALSE)", 164),
        ("NETWORKDAYS (úteis)",       "=NETWORKDAYS(F1,F2)",  "=NETWORKDAYS(F1,F2)",   120),
        ("WORKDAY +5 dias úteis",     "=WORKDAY(F1,5)",       "=WORKDAY(F1,5)",        (datetime.date(2024,1,8)-datetime.date(1899,12,30)).days),
        ("EDATE +3 meses",            "=EDATE(F1,3)",         "=EDATE(F1,3)",          (datetime.date(2024,4,1)-datetime.date(1899,12,30)).days),
        ("EOMONTH +1 mês",            "=EOMONTH(F1,1)",       "=EOMONTH(F1,1)",        (datetime.date(2024,2,29)-datetime.date(1899,12,30)).days),
        ("WEEKDAY (domingo=1)",       "=WEEKDAY(F1,1)",       "=WEEKDAY(F1,1)",        2),  # Jan 1 2024 = segunda
        ("ISOWEEKNUM",                "=ISOWEEKNUM(F1)",      "=ISOWEEKNUM(F1)",       1),
        ("TIME(12,30,0)",             "=TIME(12,30,0)",       "=TIME(12,30,0)",        0.520833333),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba Text ───────────────────────────────────────────────────────────────

def _build_text(wb: Workbook):
    ws = wb.create_sheet("Text")
    _header(ws, "Text", "713F12")
    ws["F1"] = "  Hello World  "
    ws["F2"] = "excel"
    ws["F3"] = "REVERSE"
    ws["F4"] = "2024-01-15"
    ws["F5"] = "abc123def"

    cases = [
        ("LEN (comprimento)",         "=LEN(F1)",             "=LEN(F1)",             15),
        ("TRIM (remove espaços)",      "=TRIM(F1)",            "=TRIM(F1)",            "Hello World"),
        ("UPPER (maiúsculas)",         "=UPPER(F2)",           "=UPPER(F2)",           "EXCEL"),
        ("LOWER (minúsculas)",         "=LOWER(F3)",           "=LOWER(F3)",           "reverse"),
        ("PROPER (capitaliza)",        "=PROPER(F2)",          "=PROPER(F2)",          "Excel"),
        ("LEFT(5)",                   "=LEFT(F1,5)",          "=LEFT(F1,5)",          "  Hel"),
        ("RIGHT(5)",                  "=RIGHT(F1,5)",         "=RIGHT(F1,5)",         "rld  "),
        ("MID(3,5)",                  "=MID(\"abcdefgh\",3,5)","=MID(\"abcdefgh\",3,5)","cdefg"),
        ("FIND posição",              "=FIND(\"World\",TRIM(F1))","=FIND(\"World\",TRIM(F1))", 7),
        ("SUBSTITUTE",                "=SUBSTITUTE(F2,\"l\",\"L\")","=SUBSTITUTE(F2,\"l\",\"L\")", "exceL"),
        ("REPLACE(1,5,\"---\")",      "=REPLACE(F2,1,3,\"---\")","=REPLACE(F2,1,3,\"---\")", "---el"),
        ("TEXT número como data",     "=TEXT(45292,\"DD/MM/YYYY\")","=TEXT(45292,\"DD/MM/YYYY\")","01/01/2024"),
        ("CONCATENATE",               "=CONCATENATE(\"A\",\"B\",\"C\")","=CONCATENATE(\"A\",\"B\",\"C\")", "ABC"),
        ("REPT",                      "=REPT(\"ab\",3)",      "=REPT(\"ab\",3)",      "ababab"),
        ("VALUE (texto→número)",       "=VALUE(\"3.14\")",     "=VALUE(\"3.14\")",     3.14),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba Ranking ────────────────────────────────────────────────────────────

def _build_ranking(wb: Workbook):
    ws = wb.create_sheet("Ranking")
    _header(ws, "Ranking", "831843")
    for i, v in enumerate([80, 95, 70, 95, 60], start=1):
        ws.cell(row=i, column=6, value=v)

    cases = [
        ("RANK desc (maior = 1)",     "=RANK(F1,F1:F5,0)",    "=RANK(F1,F1:F5,0)",   3),
        ("RANK asc  (menor = 1)",     "=RANK(F5,F1:F5,1)",    "=RANK(F5,F1:F5,1)",   1),
        ("RANK.EQ empate",            "=RANK.EQ(F2,F1:F5,0)", "=RANK.EQ(F2,F1:F5,0)",1),
        ("LARGE 1° maior",            "=LARGE(F1:F5,1)",      "=LARGE(F1:F5,1)",     95),
        ("LARGE 3° maior",            "=LARGE(F1:F5,3)",      "=LARGE(F1:F5,3)",     80),
        ("SMALL 1° menor",            "=SMALL(F1:F5,1)",      "=SMALL(F1:F5,1)",     60),
        ("SMALL 2° menor",            "=SMALL(F1:F5,2)",      "=SMALL(F1:F5,2)",     70),
        ("PERCENTILE 50%",            "=PERCENTILE(F1:F5,0.5)","=PERCENTILE(F1:F5,0.5)", 80.0),
        ("QUARTILE Q1",               "=QUARTILE(F1:F5,1)",   "=QUARTILE(F1:F5,1)",  70.0),
        ("MEDIAN",                    "=MEDIAN(F1:F5)",       "=MEDIAN(F1:F5)",       80.0),
        ("MODE mais frequente",       "=MODE(F1:F5)",         "=MODE(F1:F5)",         95),
        ("STDEV.S desvio padrão",     "=STDEV(F1:F5)",        "=STDEV(F1:F5)",        round(math.sqrt(sum((x-80)**2 for x in [80,95,70,95,60])/4), 6)),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba Threshold ──────────────────────────────────────────────────────────

def _build_threshold(wb: Workbook):
    ws = wb.create_sheet("Threshold")
    _header(ws, "Threshold", "134E4A")
    vals  = [100, 200, 150, 300, 50]
    cats  = ["A", "B", "A", "B", "A"]
    for i, (v, c) in enumerate(zip(vals, cats), start=1):
        ws.cell(row=i, column=6, value=v)
        ws.cell(row=i, column=7, value=c)

    maxifs_a = max(v for v, c in zip(vals, cats) if c == "A")
    minifs_b = min(v for v, c in zip(vals, cats) if c == "B")

    cases = [
        ("MAX de range",              "=MAX(F1:F5)",              "=MAX(F1:F5)",              300),
        ("MIN de range",              "=MIN(F1:F5)",              "=MIN(F1:F5)",              50),
        ("MAXIFS categoria A",        "=MAXIFS(F1:F5,G1:G5,\"A\")","=MAXIFS(F1:F5,G1:G5,\"A\")", maxifs_a),
        ("MINIFS categoria B",        "=MINIFS(F1:F5,G1:G5,\"B\")","=MINIFS(F1:F5,G1:G5,\"B\")", minifs_b),
        ("IF como limiar (>200)",     "=IF(MAX(F1:F5)>200,\"Alto\",\"Baixo\")",
                                      "=IF(MAX(F1:F5)>200,\"Alto\",\"Baixo\")", "Alto"),
        ("COUNTIF acima de 150",      "=COUNTIF(F1:F5,\">150\")", "=COUNTIF(F1:F5,\">150\")", 2),
        ("SUMIF abaixo de 200",       "=SUMIF(F1:F5,\"<200\",F1:F5)","=SUMIF(F1:F5,\"<200\",F1:F5)", 100+150+50),
        ("ISBLANK célula vazia",      "=ISBLANK(F6)",             "=ISBLANK(F6)",             True),
        ("ISNUMBER",                  "=ISNUMBER(F1)",            "=ISNUMBER(F1)",            True),
        ("ISTEXT",                    "=ISTEXT(G1)",              "=ISTEXT(G1)",              True),
        ("ISEVEN",                    "=ISEVEN(F1)",              "=ISEVEN(F1)",              True),
        ("ISODD",                     "=ISODD(F5)",               "=ISODD(F5)",               False),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba Financial ──────────────────────────────────────────────────────────

def _build_financial(wb: Workbook):
    ws = wb.create_sheet("Financial")
    _header(ws, "Financial", "1E3A5F")

    # Valores pré-calculados (fórmulas com % corretas)
    rate_m  = 0.05 / 12   # 5% a.a. / 12
    rate_6m = 0.06 / 12   # 6% a.a. / 12
    pmt_val = round(-(rate_m * (-10000) / (1 - (1+rate_m)**(-60))), 3)  # ≈ 188.712
    fv_val  = round(100 * ((1+rate_m)**12 - 1) / rate_m, 2)             # ≈ 1227.89
    pv_val  = round(-(188.71 * (1-(1+rate_m)**(-60)) / rate_m), 2)     # ≈ -9999.9
    npv_val = round(sum(v/(1.1)**(i+1) for i,v in enumerate([-1000,300,400,500])), 2)
    eff_val = round((1 + 0.10/12)**12 - 1, 5)

    cases = [
        ("PMT parcela mensal 5%/12",  "=PMT(5%/12,60,-10000)", "=PMT(5%/12,60,-10000)",  round(pmt_val, 3)),
        ("FV valor futuro 5%/12",     "=FV(5%/12,12,-100)",    "=FV(5%/12,12,-100)",     round(fv_val, 2)),
        ("PV valor presente 5%/12",   "=PV(5%/12,60,188.71)",  "=PV(5%/12,60,188.71)",   round(pv_val, 2)),
        ("NPER períodos até quitação","=NPER(5%/12,-188.71,10000)","=NPER(5%/12,-188.71,10000)", 60.0),
        ("NPV valor presente líquido","=NPV(10%,-1000,300,400,500)","=NPV(10%,-1000,300,400,500)", round(npv_val,2)),
        ("SLN depreciação linear",    "=SLN(10000,1000,5)",    "=SLN(10000,1000,5)",      1800.0),
        ("EFFECT taxa efetiva anual", "=EFFECT(10%,12)",       "=EFFECT(10%,12)",         round(eff_val, 5)),
        ("RRI taxa cresc. equivalente","=RRI(10,1000,2000)",   "=RRI(10,1000,2000)",      round(2**(1/10)-1, 5)),
        ("PMT financ. 6%/12 360x",   "=PMT(6%/12,360,-200000)","=PMT(6%/12,360,-200000)", round(rate_6m*200000/(1-(1+rate_6m)**-360), 2)),
        ("PDURATION dobrar capital",  "=PDURATION(5%,1000,2000)","=PDURATION(5%,1000,2000)", round(math.log(2)/math.log(1.05), 3)),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba Engineering ────────────────────────────────────────────────────────

def _build_engineering(wb: Workbook):
    ws = wb.create_sheet("Engineering")
    _header(ws, "Engineering", "3B0764")

    cases = [
        ("BIN2DEC \"1010\"",          "=BIN2DEC(\"1010\")",   "=BIN2DEC(\"1010\")",   10),
        ("DEC2BIN 10",                "=DEC2BIN(10)",         "=DEC2BIN(10)",         "1010"),
        ("HEX2DEC \"FF\"",            "=HEX2DEC(\"FF\")",     "=HEX2DEC(\"FF\")",     255),
        ("DEC2HEX 255",               "=DEC2HEX(255)",        "=DEC2HEX(255)",        "FF"),
        ("OCT2DEC \"17\"",            "=OCT2DEC(\"17\")",     "=OCT2DEC(\"17\")",     15),
        ("DEC2OCT 15",                "=DEC2OCT(15)",         "=DEC2OCT(15)",         "17"),
        ("BIN2HEX \"11111111\"",      "=BIN2HEX(\"11111111\")","=BIN2HEX(\"11111111\")", "FF"),
        ("BITAND 12 AND 10",          "=BITAND(12,10)",       "=BITAND(12,10)",       8),
        ("BITOR 12 OR 10",            "=BITOR(12,10)",        "=BITOR(12,10)",        14),
        ("BITXOR 12 XOR 10",          "=BITXOR(12,10)",       "=BITXOR(12,10)",       6),
        ("BITLSHIFT 1<<3",            "=BITLSHIFT(1,3)",      "=BITLSHIFT(1,3)",      8),
        ("CONVERT km→m",              "=CONVERT(1,\"km\",\"m\")","=CONVERT(1,\"km\",\"m\")", 1000.0),
        ("CONVERT °C→°F (0°C=32°F)", "=CONVERT(0,\"C\",\"F\")","=CONVERT(0,\"C\",\"F\")", 32.0),
        ("DELTA igual",               "=DELTA(5,5)",          "=DELTA(5,5)",          1),
        ("DELTA diferente",           "=DELTA(5,6)",          "=DELTA(5,6)",          0),
        ("ERF(1)",                    "=ERF(1)",              "=ERF(1)",              round(math.erf(1), 6)),
        ("IMABS \"3+4i\"",            "=IMABS(\"3+4i\")",     "=IMABS(\"3+4i\")",    5.0),
    ]
    for i, (desc, fstr, fexcel, expected) in enumerate(cases, start=2):
        _row(ws, i, desc, fstr, fexcel, expected)


# ── Aba ExternalRef (documentação) ─────────────────────────────────────────

def _build_external_ref(wb: Workbook):
    ws = wb.create_sheet("ExternalRef")
    _header(ws, "ExternalRef — Documentação", "374151")
    doc_cases = [
        ("Ref externa simples [skipped]",
         "=[Sales.xlsx]Sheet1!A1",
         "=[Sales.xlsx]Sheet1!A1",
         "#EXT!"),
        ("Ref externa com SUM [skipped]",
         "=SUM('[Budget.xlsx]Jan'!A:A)",
         "=SUM('[Budget.xlsx]Jan'!A:A)",
         "#EXT!"),
        ("IFERROR envolvendo ref ext",
         "=IFERROR('[Data.xlsx]Base'!B2, 0)",
         "=IFERROR('[Data.xlsx]Base'!B2, 0)",
         "#EXT!"),
    ]
    note_row = len(doc_cases) + 3
    for i, (desc, fstr, _, expected) in enumerate(doc_cases, start=2):
        # Gravar como string (não como fórmula) — refs externas não resolvíveis
        _row(ws, i, desc, fstr, f"'{fstr}", expected)
    ws.cell(row=note_row, column=1).value = (
        "⚠️  Células com referências externas ([workbook.xlsx]) são classificadas "
        "como EXTERNAL_REF pelo pipeline e recebem valor '#EXT!' no state_memory. "
        "Elas NÃO são avaliadas pelo formula_evaluator."
    )
    ws.cell(row=note_row, column=1).font = Font(italic=True, color="92400E", size=10)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)


# ── Orquestrador ───────────────────────────────────────────────────────────

def create_workbook() -> Path:
    wb = Workbook()
    # Remover aba default
    wb.remove(wb.active)

    _build_data(wb)
    _build_arithmetic(wb)
    _build_conditional(wb)
    _build_aggregation(wb)
    _build_lookup(wb)
    _build_date(wb)
    _build_text(wb)
    _build_ranking(wb)
    _build_threshold(wb)
    _build_financial(wb)
    _build_engineering(wb)
    _build_external_ref(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = create_workbook()
    # Contar fórmulas por aba
    from openpyxl import load_workbook as _lw
    wb2 = _lw(path)
    total = 0
    for ws in wb2.worksheets:
        count = sum(
            1 for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        )
        if count:
            print(f"  {ws.title:20s}: {count:3d} fórmulas")
            total += count
    print(f"\n✅ Planilha gerada: {path}")
    print(f"   Total de fórmulas: {total}")
    print(f"   Abas: {len(wb2.worksheets)}")
