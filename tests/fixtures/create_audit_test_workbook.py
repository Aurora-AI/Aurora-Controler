"""
Gerador da planilha de teste do Data Oracle — EXRS `audit`

Cria `sales_history_test.xlsx` com 24 meses (2023-01 a 2024-12) de histórico de vendas,
formatado exatamente como uma PME brasileira exportaria do Excel (moeda suja em texto,
datas em texto DD/MM/YYYY), com 4 anomalias PLANTADAS deliberadamente para os testes:

  1. QUEDA DE RECEITA (>3σ): mês 2024-08 tem faturamento total forçado a 15% do esperado
     pela tendência — um vazamento de receita óbvio e isolado.
  2. CHURN CRÔNICO: "Cliente Alpha Ltda" compra mensalmente (cadência ~30 dias) de
     2023-01 a 2024-06 (18 meses) e depois PARA por completo — 6 meses de silêncio,
     3x sua cadência média, sem nunca "cancelar" formalmente.
  3. PRODUTO DESCOLADO: "Adaptador USB" tem receita DECRESCENTE mês a mês enquanto a
     empresa como um todo cresce — e para de vender inteiramente a partir de 2024-09.
  4. MOEDA SUJA: toda a coluna de valor é gravada como texto "R$ X.XXX,XX" (separador de
     milhar '.', decimal ','), nunca como número — testa a coerção agressiva do mapper.

Também gera `mixed_dates_test.csv`: uma coluna de data com formatos DD/MM e MM/DD
genuinamente MISTOS na mesma coluna (não apenas ambíguos) — algumas linhas só são válidas
como DD/MM (dia > 12 na 1ª posição), outras só como MM/DD (dia > 12 na 2ª posição). Isso é
um sinal estrutural inequívoco de mistura de formatos, usado para testar a falha-alto de
ambiguidade de data no column_mapper.

Todos os valores são determinísticos (sem aleatoriedade) para que os testes possam
assertar números exatos.
"""
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "libs" / "trustware"))

try:
    from openpyxl import Workbook
except ImportError as e:
    print(f"Erro: {e}. Instale openpyxl: pip install openpyxl")
    sys.exit(1)

OUTPUT_XLSX = Path(__file__).parent / "sales_history_test.xlsx"
OUTPUT_CSV = Path(__file__).parent / "mixed_dates_test.csv"

# ── Modelo determinístico de 24 meses (2023-01 .. 2024-12) ─────────────────────────
MONTHS = [(2023 + (m // 12), (m % 12) + 1) for m in range(24)]  # (ano, mes)
LEAK_MONTH_INDEX = 19  # m=19 -> (2024, 8) = 2024-08, o mês do vazamento de receita
CHURN_LAST_ACTIVE_MONTH_INDEX = 17  # Alpha compra em m=0..17, silencia em m=18..23

PRODUCTS = ["Cabo HDMI", "Adaptador USB", "Filtro de Ar"]
CUSTOMERS = ["Cliente Alpha Ltda", "Cliente Beta SA", "Cliente Gama ME", "Cliente Delta EPP"]

# Faturamento base por produto por mês (antes do vazamento e antes do produto descolado parar).
def _cabo_hdmi(m: int) -> float:
    return 10_000.0 + m * 200.0  # cresce establemente


def _adaptador_usb(m: int) -> float:
    if m >= 20:  # para de vender inteiramente a partir de 2024-09 (m=20)
        return 0.0
    return max(5_000.0 - m * 220.0, 500.0)  # decresce enquanto a empresa cresce


def _filtro_ar(m: int) -> float:
    return 3_000.0 + m * 150.0  # cresce em linha com a empresa


_PRODUCT_FUNCS = {
    "Cabo HDMI": _cabo_hdmi,
    "Adaptador USB": _adaptador_usb,
    "Filtro de Ar": _filtro_ar,
}

# Split fixo de cada mês entre clientes (Alpha reduz a zero após CHURN_LAST_ACTIVE_MONTH_INDEX).
_CUSTOMER_SHARE_ACTIVE = {
    "Cliente Alpha Ltda": 0.40, "Cliente Beta SA": 0.30,
    "Cliente Gama ME": 0.20, "Cliente Delta EPP": 0.10,
}
_CUSTOMER_SHARE_POST_CHURN = {
    "Cliente Alpha Ltda": 0.0, "Cliente Beta SA": 0.45,
    "Cliente Gama ME": 0.35, "Cliente Delta EPP": 0.20,
}


def _customer_share(customer: str, m: int) -> float:
    if m > CHURN_LAST_ACTIVE_MONTH_INDEX:
        return _CUSTOMER_SHARE_POST_CHURN[customer]
    return _CUSTOMER_SHARE_ACTIVE[customer]


def _format_brl(value: float) -> str:
    """Formata como moeda suja pt-BR: 'R$ 1.500,00' — separador de milhar '.', decimal ','."""
    inteiro, centavos = divmod(round(value * 100), 100)
    inteiro_str = f"{inteiro:,}".replace(",", ".")
    return f"R$ {inteiro_str},{centavos:02d}"


def _format_date_ddmmyyyy(ano: int, mes: int, dia: int = 15) -> str:
    return f"{dia:02d}/{mes:02d}/{ano}"


def build_records() -> list[dict]:
    """Gera todos os registros de venda (determinísticos). Um registro por
    (mês, produto, cliente) com valor > 0."""
    records = []
    for m, (ano, mes) in enumerate(MONTHS):
        for product in PRODUCTS:
            base = _PRODUCT_FUNCS[product](m)
            if base <= 0:
                continue
            month_total = base
            if m == LEAK_MONTH_INDEX:
                month_total *= 0.15  # VAZAMENTO DE RECEITA: força a 15% do esperado
            for customer in CUSTOMERS:
                share = _customer_share(customer, m)
                if share <= 0:
                    continue
                valor = round(month_total * share, 2)
                if valor <= 0:
                    continue
                records.append({
                    "ano": ano, "mes": mes, "produto": product,
                    "cliente": customer, "valor": valor,
                    "quantidade": max(1, int(valor // 100)),
                })
    return records


def write_workbook(records: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"
    ws.append(["Data Emissão", "Produto/Serviço", "Cliente", "Valor Líquido", "Qtd"])
    for r in records:
        ws.cell(row=ws.max_row + 1, column=1,
                 value=_format_date_ddmmyyyy(r["ano"], r["mes"]))
        ws.cell(row=ws.max_row, column=2, value=r["produto"])
        ws.cell(row=ws.max_row, column=3, value=r["cliente"])
        ws.cell(row=ws.max_row, column=4, value=_format_brl(r["valor"]))
        ws.cell(row=ws.max_row, column=5, value=r["quantidade"])
    wb.save(path)


def write_mixed_dates_csv(path: Path) -> None:
    """Coluna de data com formatos DD/MM e MM/DD genuinamente misturados (não apenas
    ambíguos): linhas com dia>12 na 1ª posição (só válidas DD/MM) e linhas com dia>12 na
    2ª posição (só válidas MM/DD) coexistem na MESMA coluna."""
    lines = ["Data,Produto,Cliente,Valor,Qtd"]
    # DD/MM inequívoco: dia=25 na primeira posição (mês=03 -> março)
    lines.append("25/03/2024,Cabo HDMI,Cliente Beta SA,\"R$ 1.000,00\",5")
    lines.append("28/03/2024,Cabo HDMI,Cliente Gama ME,\"R$ 900,00\",4")
    # MM/DD inequívoco: dia=25 na SEGUNDA posição (mês=03 na primeira -> março, dia 25)
    lines.append("03/25/2024,Filtro de Ar,Cliente Delta EPP,\"R$ 700,00\",3")
    lines.append("03/28/2024,Filtro de Ar,Cliente Beta SA,\"R$ 650,00\",3")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    records = build_records()
    write_workbook(records, OUTPUT_XLSX)
    write_mixed_dates_csv(OUTPUT_CSV)
    print(f"[OK] {OUTPUT_XLSX} — {len(records)} registros")
    print(f"[OK] {OUTPUT_CSV} — datas em formatos misturados (DD/MM e MM/DD)")
