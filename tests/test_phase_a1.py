"""
EXRS Phase A1 — Testes Unitários do Extractor
"""
import sys
from pathlib import Path

# Garantir que o repositório esteja no path
REPO_ROOT = Path(__file__).resolve().parents[1]


from kernel.phase_a1.extractor import extract_structure
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
import tempfile


def _make_workbook(sheets_config: dict, save_path: Path, named_ranges: list = None, props: dict = None) -> Path:
    wb = Workbook()
    # Remover o sheet padrão se ele não estiver no config
    default_sheet = wb.active
    
    for sheet_name, cells in sheets_config.items():
        ws = wb.create_sheet(title=sheet_name)
        for cell_ref, value in cells:
            ws[cell_ref] = value
    
    if wb.active.title == "Sheet" and "Sheet" not in sheets_config:
        wb.remove(default_sheet)
        
    if named_ranges:
        for nr in named_ranges:
            wb.defined_names.add(DefinedName(nr["name"], attr_text=nr["refers_to"]))
    if props:
        if "creator" in props:
            wb.properties.creator = props["creator"]
        if "title" in props:
            wb.properties.title = props["title"]
    wb.save(save_path)
    return save_path


def test_empty_workbook():
    """Workbook sem planilhas com dados."""
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        path = _make_workbook({"Data": []}, tp / "empty.xlsx")
        ir = extract_structure(path)
        assert len(ir.sheets) == 1
        assert ir.sheets[0].info.name == "Data"
        assert len(ir.sheets[0].cells) == 0


def test_multiple_sheets():
    """Workbook com múltiplas planilhas e dados."""
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        path = _make_workbook({
            "Sheet1": [("A1", 10), ("A2", "=A1*2")],
            "Sheet2": [("B1", "Hello"), ("B2", True)],
        }, tp / "multi.xlsx")
        ir = extract_structure(path)
        # O openpyxl pode criar um 'Sheet' extra se não removermos bem,
        # mas aqui focamos na existência das que criamos.
        sheet_names = [s.info.name for s in ir.sheets]
        assert "Sheet1" in sheet_names
        assert "Sheet2" in sheet_names
        
        s1 = next(s for s in ir.sheets if s.info.name == "Sheet1")
        assert len(s1.cells) == 2


def test_named_ranges():
    """Workbook com named ranges."""
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        path = _make_workbook(
            {"Data": [("A1", 100), ("A2", 200)]},
            tp / "named.xlsx",
            named_ranges=[
                {"name": "Total", "refers_to": "Data!A1:A2"},
                {"name": "First", "refers_to": "Data!A1"},
            ]
        )
        ir = extract_structure(path)
        assert len(ir.named_ranges) >= 2


def test_hidden_sheet():
    """Workbook com planilha oculta."""
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        wb = Workbook()
        ws = wb.active
        ws.title = "Visible"
        ws["A1"] = 42
        ws2 = wb.create_sheet("Hidden")
        ws2.sheet_state = "hidden"
        ws2["A1"] = "secret"
        path = tp / "hidden.xlsx"
        wb.save(path)
        
        ir = extract_structure(path)
        visible_sheet = next(s for s in ir.sheets if s.info.name == "Visible")
        hidden_sheet = next(s for s in ir.sheets if s.info.name == "Hidden")
        assert visible_sheet.info.state == "visible"
        assert hidden_sheet.info.state == "hidden"


def test_metadata():
    """Workbook com metadados."""
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        path = _make_workbook(
            {"Data": []},
            tp / "meta.xlsx",
            props={"creator": "Rodrigo", "title": "Test Workbook"}
        )
        ir = extract_structure(path)
        assert ir.metadata["creator"] == "Rodrigo"
        assert ir.metadata["title"] == "Test Workbook"


if __name__ == "__main__":
    try:
        test_empty_workbook()
        test_multiple_sheets()
        test_named_ranges()
        test_hidden_sheet()
        test_metadata()
        print("[TEST A1.5] Todos os testes passaram.")
    except Exception as e:
        print(f"[TEST A1.5] FALHA: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
