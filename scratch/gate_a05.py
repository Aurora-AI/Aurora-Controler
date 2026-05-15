import sys, json, tempfile
from pathlib import Path

# Setup de paths
sys.path.insert(0, 'C:/Projetos/ExcelReverseEngine/src/phase_a0')
sys.path.insert(0, 'C:/Projetos/ExcelReverseEngine/libs/trustware')

from classifier import classify_workbook
from pipeline_contracts import CompatibilityReport
from openpyxl import Workbook

with tempfile.TemporaryDirectory() as td:
    tp = Path(td)
    
    # Gerar planilha SUPPORTED de controle
    wb = Workbook()
    ws = wb.active
    ws.title = 'Control'
    ws['A1'] = 100
    ws['A2'] = 200
    ws['A3'] = '=SUM(A1:A2)'
    ws['B1'] = '=IF(A3>0, "POSITIVE", "NEGATIVE")'
    path = tp / 'gate_final.xlsx'
    wb.save(path)
    
    # Classificar
    report = classify_workbook(path)
    
    # Validar via Pydantic (dispara ValidationError se falhar)
    validated = CompatibilityReport.model_validate(report.model_dump())
    data = json.loads(validated.model_dump_json())
    
    # Asserts
    assert data['workbook_class'] == 'supported', f'Esperado supported, obtido {data["workbook_class"]}'
    assert data['compile_decision'] == 'proceed', f'Esperado proceed, obtido {data["compile_decision"]}'
    assert data['escalate_reasons'] is None
    assert data['restricted_items'] is None
    
    print('[GATE A0.5] Classificação final validada — schema íntegro e output correto.')
    print(json.dumps(data, indent=2))
