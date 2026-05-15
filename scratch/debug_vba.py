import openpyxl
from pathlib import Path
import tempfile

with tempfile.TemporaryDirectory() as td:
    path = Path(td) / "test.xlsx"
    wb = openpyxl.Workbook()
    wb.save(path)
    
    wb2 = openpyxl.load_workbook(path, keep_vba=True)
    print(f"keep_vba=True: namelist has vbaProject.bin? {'xl/vbaProject.bin' in wb2.vba_archive.namelist()}")
    wb2.close()
    
    wb3 = openpyxl.load_workbook(path, read_only=True, keep_vba=True)
    print(f"read_only=True, keep_vba=True: namelist has vbaProject.bin? {'xl/vbaProject.bin' in wb3.vba_archive.namelist()}")
    wb3.close()
