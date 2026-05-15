import sys, tempfile
from pathlib import Path

# Setup paths
REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src" / "phase_a1"))
sys.path.insert(0, str(REPO_ROOT / "libs" / "trustware"))

from extractor import extract_structure
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

with tempfile.TemporaryDirectory() as td:
    tp = Path(td)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    ws['A1'] = 10
    ws['A2'] = 20
    ws['A3'] = 30
    wb.defined_names.add(DefinedName('TotalRange', attr_text='Data!A1:A3'))
    wb.defined_names.add(DefinedName('FirstCell', attr_text='Data!A1'))
    path = tp / 'named.xlsx'
    wb.save(path)
    
    ir = extract_structure(path)
    assert len(ir.named_ranges) >= 2, f'Esperado pelo menos 2 named ranges, obtido {len(ir.named_ranges)}'
    
    total = next(nr for nr in ir.named_ranges if nr.name == 'TotalRange')
    assert 'Data' in total.refers_to
    assert total.scope is None
    
    print('[GATE A1.4] Named ranges extraídos corretamente.')
    for nr in ir.named_ranges:
        print(f'  {nr.name} -> {nr.refers_to} (scope: {nr.scope})')
