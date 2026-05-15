import sys, tempfile, json
from pathlib import Path

# Setup paths
REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src" / "phase_a1"))
sys.path.insert(0, str(REPO_ROOT / "libs" / "trustware"))

from extractor import extract_structure
from pipeline_contracts import WorkbookIR
from openpyxl import Workbook

with tempfile.TemporaryDirectory() as td:
    tp = Path(td)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Final'
    ws['A1'] = 'Name'
    ws['A2'] = 'Alice'
    ws['B1'] = 'Score'
    ws['B2'] = 95
    ws['B3'] = '=AVERAGE(B2)'
    wb.create_sheet('Config').sheet_state = 'hidden'
    wb.properties.creator = 'EXRS'
    path = tp / 'final.xlsx'
    wb.save(path)
    
    ir = extract_structure(path)
    
    # Validar via Pydantic
    # ir.model_dump() é compatível com o schema
    validated = WorkbookIR.model_validate(ir.model_dump())
    data = json.loads(validated.model_dump_json())
    
    assert len(data['sheets']) == 2
    assert data['metadata']['creator'] == 'EXRS'
    assert len(data['sheets'][0]['cells']) > 0
    
    # Verificar se a fórmula foi capturada
    cells = data['sheets'][0]['cells']
    b3 = next(c for c in cells if c['coordinate'] == 'B3')
    assert b3['formula'] == '=AVERAGE(B2)'
    
    print('[GATE A1.6] IR bruto validado — schema íntegro e extração completa.')
