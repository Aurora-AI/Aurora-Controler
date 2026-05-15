import sys, tempfile
from pathlib import Path

# Setup paths
REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src" / "phase_a1"))
sys.path.insert(0, str(REPO_ROOT / "libs" / "trustware"))

from extractor import extract_structure
from openpyxl import Workbook

with tempfile.TemporaryDirectory() as td:
    tp = Path(td)
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 100
    ws['A2'] = '=A1*2'
    ws['B1'] = 'Texto'
    ws['B2'] = True
    ws['C1'] = '=NOW()'
    path = tp / 'cells.xlsx'
    wb.save(path)
    
    ir = extract_structure(path)
    cells = ir.sheets[0].cells
    
    assert len(cells) >= 5, f'Esperado pelo menos 5 células, obtido {len(cells)}'
    
    a1 = next(c for c in cells if c.coordinate == 'A1')
    assert a1.data_type == 'n'
    assert a1.value == 100
    assert a1.formula is None
    
    a2 = next(c for c in cells if c.coordinate == 'A2')
    assert a2.formula == '=A1*2'
    assert a2.data_type == 'e'
    
    b1 = next(c for c in cells if c.coordinate == 'B1')
    assert b1.data_type == 's'
    
    b2 = next(c for c in cells if c.coordinate == 'b2'.upper())
    assert b2.data_type == 'b'
    
    c1 = next(c for c in cells if c.coordinate == 'C1')
    assert c1.formula == '=NOW()'
    
    print('[GATE A1.3] Células extraídas com tipos e fórmulas corretos.')
    print(f'Total de células: {len(cells)}')
