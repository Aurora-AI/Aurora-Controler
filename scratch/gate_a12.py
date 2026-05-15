import sys, tempfile
from pathlib import Path

# Setup paths
REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src" / "phase_a1"))
sys.path.insert(0, str(REPO_ROOT / "libs" / "trustware"))

from extractor import extract_structure
from openpyxl import Workbook

with tempfile.TemporaryDirectory() as td:
    tp = Path(td)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    ws['A1'] = 100
    ws2 = wb.create_sheet('Hidden')
    ws2.sheet_state = 'hidden'
    ws2['A1'] = 'secret'
    wb.properties.creator = 'EXRS Test'
    path = tp / 'test.xlsx'
    wb.save(path)
    
    ir = extract_structure(path)
    
    assert len(ir.sheets) == 2, f'Esperado 2 sheets, obtido {len(ir.sheets)}'
    assert ir.sheets[0].info.name == 'Data'
    assert ir.sheets[0].info.index == 0
    assert ir.sheets[0].info.state == 'visible'
    assert ir.sheets[1].info.name == 'Hidden'
    assert ir.sheets[1].info.state == 'hidden'
    assert ir.metadata['creator'] == 'EXRS Test'
    assert len(ir.sheets[0].cells) == 0  # sem células ainda (esperado na A1.2)
    
    print('[GATE A1.2] Estrutura extraída corretamente.')
