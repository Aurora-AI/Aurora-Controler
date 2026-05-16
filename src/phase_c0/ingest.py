"""Fase C0 — Leitura do arquivo de entrada (.csv / .xlsx)."""
from __future__ import annotations

from pathlib import Path
from typing import Any


def read_table(path: str) -> tuple[str, list[list[Any]]]:
    """Lê .csv ou .xlsx e retorna (nome_da_aba, linhas).

    Cada linha é uma lista de células na ordem das colunas.
    """
    p = Path(path)
    suffix = p.suffix.lower()

    if suffix == ".csv":
        import csv
        with p.open(encoding="utf-8", newline="") as fh:
            rows = [list(r) for r in csv.reader(fh)]
        return (p.stem, rows)

    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=True, read_only=True)
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        sheet = ws.title
        wb.close()
        return (sheet, rows)

    raise ValueError(f"Formato não suportado: {suffix}")
