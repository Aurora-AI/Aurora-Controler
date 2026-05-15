"""
EXRS Phase A4 — Gabarito Extractor
Lê valores pré-calculados do Excel (data_only=True) e detecta cache ausente.
"""
import sys
from pathlib import Path
from typing import Any, Tuple

# Setup paths for trustware
REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "libs" / "trustware"))

from openpyxl import load_workbook

def extract_gabarito(filepath: Path) -> Tuple[dict, set]:
    """
    Extrai valores calculados de todas as células como gabarito.
    Detecta células que têm fórmulas mas não têm valor calculado em cache.
    
    Retorna:
    - gabarito: dict[node_id, valor]
    - cells_without_cache: set de node_ids com fórmula mas sem cache
    """
    # Leitura 1: data_only=True (valores cacheados pelo engine do Excel)
    wb = load_workbook(filepath, read_only=True, data_only=True)
    gabarito: dict[str, Any] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    node_id = f"{ws.title}!{cell.coordinate}"
                    gabarito[node_id] = cell.value
    wb.close()
    
    # Leitura 2: data_only=False (para identificar onde há fórmulas mas o cache acima retornou None)
    wb_raw = load_workbook(filepath, read_only=True, data_only=False)
    cells_without_cache = set()
    for ws in wb_raw.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                # Se o valor é uma string começando com '=', é uma fórmula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    node_id = f"{ws.title}!{cell.coordinate}"
                    if node_id not in gabarito:
                        cells_without_cache.add(node_id)
    wb_raw.close()
    
    return gabarito, cells_without_cache

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="EXRS Phase A4 — Gabarito Extractor")
    parser.add_argument("input", type=Path, help="Caminho para .xlsx")
    args = parser.parse_args()
    
    if not args.input.exists():
        print(f"Erro: Arquivo não encontrado: {args.input}")
        sys.exit(1)
        
    gab, missing = extract_gabarito(args.input)
    print(f"Gabarito extraído: {len(gab)} células.")
    print(f"Células com fórmula mas sem cache: {len(missing)}")
    if missing:
        print(f"Exemplos: {list(missing)[:5]}")
