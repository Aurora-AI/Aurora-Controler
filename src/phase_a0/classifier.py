"""
EXRS Phase A0 — Unsupported Construct Classifier
Deterministic workbook complexity classifier.
No LLM. No IO except file read and JSON output.
"""
import sys
import json
import argparse
from pathlib import Path

# Garantir que libs/trustware esteja no path
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "libs" / "trustware"))

import re
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from pipeline_contracts import CompatibilityReport, WorkbookClass, CompileDecision

_EXT_REF_RE = re.compile(r'\[[^\]]+\.(xlsx|xlsm|xls|xlsb|xlam|csv)\]', re.IGNORECASE)

def _formula_has_external_ref(formula: str) -> bool:
    if not formula or '[' not in formula:
        return False
    without_strings = re.sub(r'"[^"]*"', '', formula)
    return bool(_EXT_REF_RE.search(without_strings))

# Constantes da Spec — funções voláteis que indicam lógica RESTRICTED
VOLATILE_FUNCTIONS = {
    'NOW', 'TODAY', 'RAND', 'RANDBETWEEN', 'OFFSET', 'INDIRECT',
    'CELL', 'INFO', 'RANDARRAY',
}

# Limites de amostragem para classificação (não extração)
MAX_SAMPLE_ROWS = 1000
MAX_SAMPLE_COLS = 50


def classify_formula(formula: str | None) -> str:
    """
    Classifica uma fórmula como 'SUPPORTED' ou 'RESTRICTED'.
    Usa Tokenizer para inspecionar apenas tokens de função, evitando
    falsos positivos em strings literais ou comentários.
    """
    if formula is None:
        return "SUPPORTED"
    try:
        tok = Tokenizer(formula)
        for token in tok.items:
            if token.type.startswith("FUNC"):
                func_name = token.value.rstrip("(").upper()
                if func_name in VOLATILE_FUNCTIONS:
                    return "RESTRICTED"
    except Exception:
        # Fallback: se o tokenizer falhar, checar presença simples do nome
        formula_upper = formula.upper()
        for vf in VOLATILE_FUNCTIONS:
            if f"{vf}(" in formula_upper:
                return "RESTRICTED"
    return "SUPPORTED"


def classify_workbook(filepath: Path) -> CompatibilityReport:
    """
    Classifica um workbook .xlsx quanto à elegibilidade para o pipeline EXRS.
    
    Retorna CompatibilityReport com:
    - workbook_class: SUPPORTED, RESTRICTED ou UNSUPPORTED
    - compile_decision: PROCEED, PROCEED_WITH_RESTRICTIONS ou ESCALATE
    - escalate_reasons: lista de motivos de escalação (se houver)
    - restricted_items: lista de itens restritos encontrados (se houver)
    """
    print(f"[INFO] Carregando workbook: {filepath.name}")
    
    if filepath.suffix.lower() == ".csv":
        return CompatibilityReport(
            workbook_class=WorkbookClass.SUPPORTED,
            compile_decision=CompileDecision.PROCEED,
            construct_details=[{"type": "csv_tabular", "severity": "supported", "detail": "CSV file detected"}]
        )

    try:
        wb = load_workbook(filepath, read_only=True, data_only=False, keep_vba=True)
    except Exception as e:
        return CompatibilityReport(
            workbook_class=WorkbookClass.UNSUPPORTED,
            compile_decision=CompileDecision.ESCALATE,
            escalate_reasons=[f"Failed to load workbook: {str(e)}"]
        )
    
    # Detecção real de VBA (vba_archive contém todo o zip se keep_vba=True)
    has_vba = False
    if getattr(wb, 'vba_archive', None):
        has_vba = 'xl/vbaProject.bin' in wb.vba_archive.namelist()
    
    has_external_links = False
    classification = WorkbookClass.SUPPORTED
    construct_details: list[dict] = []
    escalate_reasons: list[str] = []
    restricted_items: list[dict] = []

    # --- 1. Detecção de VBA (UNSUPPORTED) ---
    if has_vba:
        classification = WorkbookClass.UNSUPPORTED
        escalate_reasons.append("VBA/macros detected in workbook")
        construct_details.append({
            "type": "vba",
            "severity": "unsupported",
            "detail": "Workbook contains VBA project (vba_archive is not None)"
        })

    # --- 2. Detecção de links externos (UNSUPPORTED) ---
    try:
        defined_names_iter = list(wb.defined_names.values())
    except Exception:
        defined_names_iter = []
    for name in defined_names_iter:
        name_str = getattr(name, 'name', '')
        attr_text = getattr(name, 'attr_text', getattr(name, 'value', '')) or ''
        if '[' in attr_text or '[' in name_str:
            has_external_links = True
            escalate_reasons.append(f"External workbook link detected: {name_str}")
            construct_details.append({
                "type": "external_link",
                "name": name_str,
                "severity": "unsupported"
            })
    
    if has_external_links:
        classification = WorkbookClass.UNSUPPORTED

    # --- 3. Amostragem de fórmulas por planilha ---
    for ws in wb.worksheets:
        print(f"[INFO] Analisando planilha: {ws.title}...")
        sheet_restricted = False
        
        # Amostragem: até MAX_SAMPLE_ROWS linhas, MAX_SAMPLE_COLS colunas
        effective_max_row = min(ws.max_row or 0, MAX_SAMPLE_ROWS)
        effective_max_col = min(ws.max_column or 0, MAX_SAMPLE_COLS)
        
        if effective_max_row == 0 or effective_max_col == 0:
            continue
            
        for row in ws.iter_rows(
            min_row=1, max_row=effective_max_row,
            min_col=1, max_col=effective_max_col
        ):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                    # Detectar referências externas em fórmulas de célula
                    if _formula_has_external_ref(cell.value):
                        restricted_items.append({
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "formula": cell.value[:200],
                            "type": "external_ref"
                        })
                        sheet_restricted = True
                        continue

                    result = classify_formula(cell.value)
                    if result == "RESTRICTED":
                        restricted_items.append({
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "formula": cell.value[:200]
                        })
                        sheet_restricted = True
        
        # Se a planilha está oculta e contém fórmulas, marcar como restricted adicional
        if sheet_restricted and ws.sheet_state == 'hidden':
            construct_details.append({
                "type": "hidden_sheet_formula",
                "sheet": ws.title,
                "severity": "restricted",
                "detail": "Hidden sheet contains formulas; dependency chain may be obscured"
            })

    # --- 4. Determinar classificação final ---
    if classification == WorkbookClass.SUPPORTED and restricted_items:
        classification = WorkbookClass.RESTRICTED

    # --- 5. Determinar decisão de compilação ---
    if classification == WorkbookClass.UNSUPPORTED:
        compile_decision = CompileDecision.ESCALATE
    elif classification == WorkbookClass.RESTRICTED:
        compile_decision = CompileDecision.PROCEED_WITH_RESTRICTIONS
    else:
        compile_decision = CompileDecision.PROCEED

    wb.close()
    
    return CompatibilityReport(
        workbook_class=classification,
        compile_decision=compile_decision,
        escalate_reasons=escalate_reasons if escalate_reasons else None,
        restricted_items=restricted_items if restricted_items else None,
        construct_details=construct_details if construct_details else None
    )


# ─── CLI ────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="EXRS Phase A0 — Unsupported Construct Classifier"
    )
    parser.add_argument(
        "input", type=Path,
        help="Caminho para arquivo .xlsx a ser classificado"
    )
    parser.add_argument(
        "--output", type=Path, default=None,
        help="Caminho para salvar compatibility_report.json (default: <input>.compatibility_report.json)"
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Erro: Arquivo não encontrado: {args.input}")
        sys.exit(1)

    report = classify_workbook(args.input)
    output_path = args.output or args.input.with_suffix(".compatibility_report.json")
    output_path.write_text(report.model_dump_json(indent=2), encoding="utf-8")
    
    print(f"Classification: {report.workbook_class.value}")
    print(f"Decision: {report.compile_decision.value}")
    print(f"Report saved to: {output_path}")
