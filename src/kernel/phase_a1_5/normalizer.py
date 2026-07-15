"""
EXRS Phase A1.5 — Canonical IR Normalizer
Normalização determinística de fórmulas e entidades.
"""
import sys
import re
import logging
from pathlib import Path
from typing import List, Dict, Optional, Any

_log = logging.getLogger(__name__)

from openpyxl.formula import Tokenizer
from openpyxl.utils import get_column_letter, column_index_from_string
from kernel.contracts import (
    WorkbookIR, NormalizedWorkbookIR, NormalizedSheet, NormalizedCell,
    FormulaToken, FormulaTokenType, CellData, SheetData
)

# Mapeamento de funções pt-BR → en-US (canônico)
PT_BR_TO_EN: Dict[str, str] = {
    'SOMA': 'SUM', 'MÉDIA': 'AVERAGE', 'MÉD': 'AVERAGE',
    'SE': 'IF', 'PROCV': 'VLOOKUP', 'PROCH': 'HLOOKUP',
    'CORRESP': 'MATCH', 'ÍNDICE': 'INDEX', 'DESLOC': 'OFFSET',
    'CONT.SE': 'COUNTIF', 'CONT.SES': 'COUNTIFS',
    'SOMASE': 'SUMIF', 'SOMASES': 'SUMIFS',
    'HOJE': 'TODAY', 'AGORA': 'NOW', 'ALEATÓRIO': 'RAND',
    'ESQUERDA': 'LEFT', 'DIREITA': 'RIGHT',
    'MAIÚSCULA': 'UPPER', 'MINÚSCULA': 'LOWER',
    'TEXTO': 'TEXT', 'VALOR': 'VALUE',
    'ÉCÉL.VAZIA': 'ISBLANK', 'ÉERRO': 'ISERROR',
    'ARRED': 'ROUND', 'INT': 'INT',
    'OU': 'OR', 'E': 'AND', 'NÃO': 'NOT',
    'DATA': 'DATE', 'ANO': 'YEAR', 'MÊS': 'MONTH', 'DIA': 'DAY',
}

FULL_COLUMN_PATTERN = re.compile(r'^([A-Z]+):([A-Z]+)$', re.IGNORECASE)

# Detecta referência a workbook externo: [arquivo.xlsx] fora de strings literais
_EXT_REF_RE = re.compile(r'\[[^\]]+\.(xlsx|xlsm|xls|xlsb|xlam|csv)\]', re.IGNORECASE)

def _has_external_ref(formula: str) -> bool:
    """Retorna True se a fórmula contém [workbook.xlsx] fora de strings literais."""
    if not formula or '[' not in formula:
        return False
    without_strings = re.sub(r'"[^"]*"', '', formula)
    return bool(_EXT_REF_RE.search(without_strings))

def detect_locale(raw_ir: WorkbookIR) -> str:
    """Detecta locale predominante do workbook."""
    pt_br_count = 0
    en_count = 0
    for sheet_data in raw_ir.sheets:
        for cell in sheet_data.cells:
            if not cell.formula:
                continue
            formula_upper = cell.formula.upper()
            for pt_func in PT_BR_TO_EN:
                if f'{pt_func}(' in formula_upper:
                    pt_br_count += 1
            # Se não encontrou pt-BR e tem funções comuns em inglês
            en_funcs = ['SUM(', 'IF(', 'VLOOKUP(', 'AVERAGE(', 'COUNT(']
            for en_func in en_funcs:
                if en_func in formula_upper:
                    en_count += 1
    return 'pt-BR' if pt_br_count > en_count else 'en-US'

def translate_function_name(name: str, from_locale: str) -> str:
    """Traduz nome de função para inglês canônico."""
    if from_locale == 'pt-BR':
        return PT_BR_TO_EN.get(name.upper(), name)
    return name

def _map_token_type(openpyxl_type_str) -> FormulaTokenType:
    """Mapeia tipo do openpyxl Token para FormulaTokenType."""
    t = openpyxl_type_str.upper()
    if 'OPERAND' in t: return FormulaTokenType.OPERAND
    if 'FUNC' in t: return FormulaTokenType.FUNCTION
    if 'OPEN' in t: return FormulaTokenType.LPAREN
    if 'CLOSE' in t: return FormulaTokenType.RPAREN
    if 'SEP' in t: return FormulaTokenType.SEPARATOR
    if 'OP' in t: return FormulaTokenType.OPERATOR
    if 'NUM' in t: return FormulaTokenType.CONSTANT
    if 'STR' in t: return FormulaTokenType.CONSTANT
    if 'LITERAL' in t: return FormulaTokenType.CONSTANT
    return FormulaTokenType.CONSTANT

def tokenize_formula(formula: str, locale: str = 'en-US') -> list[FormulaToken]:
    """Tokeniza uma fórmula Excel usando openpyxl Tokenizer."""
    if not formula or not formula.startswith('='):
        return []
    
    try:
        tok = Tokenizer(formula)
        tokens: list[FormulaToken] = []
        for item in tok.items:
            token_type = _map_token_type(item.type)
            value = item.value
            
            # Ajuste de tipos baseado no valor se o Tokenizer falhar na classificação fina
            if token_type == FormulaTokenType.CONSTANT and re.match(r'^[A-Z]+\d+', value.upper()):
                 token_type = FormulaTokenType.OPERAND
            
            if token_type == FormulaTokenType.FUNCTION:
                if value.endswith('('):
                    value = value[:-1]
                elif value == ')':
                    token_type = FormulaTokenType.RPAREN
            
            if token_type == FormulaTokenType.FUNCTION and locale != 'en-US':
                value = translate_function_name(value, locale)
            
            tokens.append(FormulaToken(
                type=token_type,
                value=value,
                position=item.position if hasattr(item, 'position') else 0
            ))
        return tokens
    except Exception as exc:
        _log.debug(
            "tokenize_formula: falha ao tokenizar %r (%s) — usando token degenerado",
            formula[:80], exc
        )
        return [FormulaToken(
            type=FormulaTokenType.FUNCTION,
            value=formula[1:] if formula.startswith('=') else formula,
            position=0
        )]

def normalize_value(value, data_type: str, locale: str = 'en-US'):
    """Normaliza valor para tipo Python canônico."""
    if value is None:
        return None
    
    if data_type == 'n':
        if isinstance(value, str):
            if locale == 'pt-BR':
                value = value.replace('.', '').replace(',', '.')
            try:
                return float(value)
            except ValueError:
                return value
        return value
    
    elif data_type == 'b':
        if isinstance(value, str):
            return value.upper() == 'TRUE' or value == '1'
        return bool(value)
    
    elif data_type == 's':
        return str(value)
    
    elif data_type == 'e':
        return str(value) if value else None
    
    return value

def is_null_value(value, data_type: str) -> bool:
    """Determina se o valor é considerado nulo para fins de análise."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    # Erros do Excel são valores, não nulos
    if data_type == 'e':
        return False
    return False

def detect_merged_cells(ws) -> Dict[str, str]:
    """Mapeia células filhas -> célula mestre em ranges merged."""
    merged_map: Dict[str, str] = {}
    if not hasattr(ws, 'merged_cells'):
        return merged_map
        
    for merged_range in ws.merged_cells.ranges:
        # A célula mestre é top-left
        master = merged_range.min_row, merged_range.min_col
        master_coord = f"{get_column_letter(master[1])}{master[0]}"
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                if coord != master_coord:
                    merged_map[coord] = master_coord
    return merged_map

def parse_cell_coordinate(coord: str) -> tuple[str, int]:
    """Separa 'A1' em ('A', 1)."""
    match = re.match(r'^([A-Z]+)(\d+)$', coord.upper().replace('$', ''))
    if not match:
        raise ValueError(f"Coordenada inválida: {coord}")
    return match.group(1), int(match.group(2))

def expand_range(range_str: str, max_row: int | None = None) -> list[str]:
    """Expande um range como 'A1:B2' em lista de coordenadas individuais."""
    range_str = range_str.replace('$', '')
    
    full_col_match = FULL_COLUMN_PATTERN.match(range_str)
    if full_col_match:
        if max_row is None:
            max_row = 2000
        col1 = column_index_from_string(full_col_match.group(1))
        col2 = column_index_from_string(full_col_match.group(2))
        start_col = min(col1, col2)
        end_col = max(col1, col2)
        cells: list[str] = []
        for row in range(1, max_row + 1):
            for col in range(start_col, end_col + 1):
                cells.append(f"{get_column_letter(col)}{row}")
        return cells
    
    cells: list[str] = []
    if ':' in range_str:
        parts = range_str.split(':')
        if len(parts) != 2:
            return cells
        try:
            col1, row1 = parse_cell_coordinate(parts[0])
            col2, row2 = parse_cell_coordinate(parts[1])
        except ValueError:
            return cells
        
        col1_idx = column_index_from_string(col1)
        col2_idx = column_index_from_string(col2)
        start_col = min(col1_idx, col2_idx)
        end_col = max(col1_idx, col2_idx)
        start_row = min(row1, row2)
        end_row = max(row1, row2)
        
        if end_row - start_row > 2000:
            end_row = start_row + 2000
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cells.append(f"{get_column_letter(col)}{row}")
    else:
        cells.append(range_str)
    
    return cells

def resolve_sheet_reference(ref: str, current_sheet: str) -> tuple[str, str]:
    """Resolve referência que pode incluir sheet: 'Sheet2!A1' -> ('Sheet2', 'A1')."""
    if '!' in ref:
        sheet_part, cell_part = ref.split('!', 1)
        sheet_part = sheet_part.strip("'\"")
        return sheet_part, cell_part
    return current_sheet, ref

def extract_dependencies(tokens: list[FormulaToken], current_sheet: str, max_row: int | None = None) -> list[str]:
    """Extrai lista de coordenadas que esta fórmula referencia."""
    dependencies: set[str] = set()
    for token in tokens:
        if token.type == FormulaTokenType.OPERAND:
            token_val = token.value.replace('$', '')
            if _looks_like_cell_ref(token_val):
                try:
                    sheet, cell_ref = resolve_sheet_reference(token_val, current_sheet)
                    expanded = expand_range(cell_ref, max_row=max_row)
                    for cell in expanded:
                        full_ref = f"{sheet}!{cell}" if sheet != current_sheet else cell
                        dependencies.add(full_ref)
                except (ValueError, IndexError) as exc:
                    _log.debug(
                        "extract_dependencies: não resolveu referência %r em %r (%s)",
                        token_val, current_sheet, exc
                    )
    return sorted(dependencies)

def _looks_like_cell_ref(value: str) -> bool:
    """Heurística: verifica se o valor parece uma referência de célula ou range."""
    cell_part = value.split('!')[-1] if '!' in value else value
    cell_part = cell_part.upper().replace('$', '')
    # Full-column reference: N:N or A:Z
    if re.match(r'^[A-Z]+:[A-Z]+$', cell_part):
        return True
    return bool(re.match(r'^[A-Z]+\d+', cell_part))

def normalize_workbook(raw_ir: WorkbookIR) -> NormalizedWorkbookIR:
    """Orquestra a normalização completa do workbook."""
    locale = detect_locale(raw_ir)
    normalized_sheets: List[NormalizedSheet] = []
    
    for sheet_data in raw_ir.sheets:
        norm_cells: List[NormalizedCell] = []
        
        # Determinar max_row para esta planilha para expansão de A:A
        rows = []
        for c in sheet_data.cells:
            try:
                _, r = parse_cell_coordinate(c.coordinate)
                rows.append(r)
            except Exception as exc:
                _log.debug(
                    "normalize_workbook: coordenada inválida %r (%s)",
                    c.coordinate, exc
                )
        
        max_row = max(rows) if rows else 2000
        
        for cell in sheet_data.cells:
            has_ext = _has_external_ref(cell.formula)
            tokens = tokenize_formula(cell.formula, locale) if cell.formula and not has_ext else []
            deps = extract_dependencies(tokens, sheet_data.info.name, max_row=max_row)

            norm_val = normalize_value(cell.value, cell.data_type, locale)

            norm_cells.append(NormalizedCell(
                coordinate=cell.coordinate,
                value_static=norm_val,
                formula_raw=cell.formula,
                formula_tokens=tokens,
                dependencies=deps,
                data_type=cell.data_type,
                locale_original=locale,
                has_external_ref=has_ext,
            ))
            
        normalized_sheets.append(NormalizedSheet(
            name=sheet_data.info.name,
            index=sheet_data.info.index,
            state=sheet_data.info.state,
            cells=norm_cells
        ))
        
    return NormalizedWorkbookIR(
        file_path=raw_ir.file_path,
        sheets=normalized_sheets,
        named_ranges=raw_ir.named_ranges,
        metadata=raw_ir.metadata,
        normalization_info={'locale': locale}
    )
