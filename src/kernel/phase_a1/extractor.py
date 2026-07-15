"""
EXRS Phase A1 — Structural Extractor
Lê Excel como OpenXML package — data_only=False para capturar fórmulas.
Determinístico. Sem LLM.
"""
import sys
from pathlib import Path

from openpyxl import load_workbook
from kernel.contracts import WorkbookIR, SheetData, SheetInfo, CellData, NamedRangeInfo


# Constantes de extração
MAX_EXTRACT_ROWS = 2000
MAX_EXTRACT_COLS = 100


def extract_cells(ws) -> list[CellData]:
    """Extrai células não vazias de uma planilha (data_only=False)."""
    cells: list[CellData] = []
    
    # Em modo read_only, max_row/max_column podem ser None
    m_row = ws.max_row if ws.max_row else MAX_EXTRACT_ROWS
    m_col = ws.max_column if ws.max_column else MAX_EXTRACT_COLS
    
    effective_max_row = min(m_row, MAX_EXTRACT_ROWS)
    effective_max_col = min(m_col, MAX_EXTRACT_COLS)
    
    for row in ws.iter_rows(
        min_row=1, max_row=effective_max_row,
        min_col=1, max_col=effective_max_col
    ):
        for cell in row:
            if cell.value is None:
                continue
            
            # Determinar data_type
            dt = getattr(cell, 'data_type', 's')
            
            # Extrair fórmula se existir
            formula = None
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                dt = 'e'  # forçar tipo fórmula se começar com =
            
            cells.append(CellData(
                coordinate=cell.coordinate,
                value=cell.value,
                formula=formula,
                data_type=dt
            ))
    
    return cells


def extract_structure(filepath: Path) -> WorkbookIR:
    """Extrai metadados e estrutura do workbook — data_only=False."""
    # load_workbook com data_only=False captura as fórmulas
    wb = load_workbook(filepath, read_only=True, data_only=False)
    
    try:
        # Metadados do workbook
        metadata = {
            "creator": wb.properties.creator or "",
            "created": str(wb.properties.created) if wb.properties.created else "",
            "modified": str(wb.properties.modified) if wb.properties.modified else "",
            "last_modified_by": wb.properties.lastModifiedBy or "",
            "title": wb.properties.title or "",
            "subject": wb.properties.subject or "",
            "category": wb.properties.category or "",
        }
        
        sheets: list[SheetData] = []
        for idx, ws in enumerate(wb.worksheets):
            # No modo read_only, alguns atributos podem não estar presentes
            # se a planilha for muito grande ou não tiver metadados de dimensão.
            dimensions = None
            if hasattr(ws, 'dimensions'):
                dimensions = ws.dimensions
            
            info = SheetInfo(
                name=ws.title,
                index=idx,
                state=ws.sheet_state or "visible",
                dimensions=dimensions,
                max_row=getattr(ws, 'max_row', None),
                max_column=getattr(ws, 'max_column', None),
            )
            
            # Extração de células (Micro Etapa A1.3)
            cells = extract_cells(ws)
            sheets.append(SheetData(info=info, cells=cells))
        
        # Extração de Named Ranges (Micro Etapa A1.4)
        named_ranges: list[NamedRangeInfo] = []
        if wb.defined_names is not None:
            # Em versões recentes do openpyxl, defined_names é um objeto iterável
            # que contém instâncias de DefinedName.
            try:
                # Tentar iterar diretamente (comum em read_only)
                for name, dn in wb.defined_names.items():
                    scope = None
                    if hasattr(dn, 'localSheetId') and dn.localSheetId is not None:
                        try:
                            scope = wb.worksheets[dn.localSheetId].title
                        except (IndexError, KeyError):
                            scope = f"sheet_id_{dn.localSheetId}"
                    
                    named_ranges.append(NamedRangeInfo(
                        name=name,
                        refers_to=getattr(dn, 'attr_text', '') or '',
                        scope=scope
                    ))
            except AttributeError:
                # Fallback para caso seja uma lista em versões específicas
                pass

        return WorkbookIR(
            file_path=str(filepath.resolve()),
            sheets=sheets,
            named_ranges=named_ranges,
            metadata=metadata,
        )
    finally:
        wb.close()
