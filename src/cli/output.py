"""
EXRS CLI — Montagem da pasta de saída limpa.

Por padrão entrega só o .py (+ engine vendorizado) e o relatório .html — os JSONs técnicos
por fase continuam existindo em job_output_dir (nada se perde do rastro determinístico),
mas só são copiados para a pasta final de entrega quando debug=True.
"""
import shutil
from pathlib import Path

from product_a.phase_a4.html_reporter import generate_html_report

from cli.codegen import render_replay_module

_REPO_ROOT = Path(__file__).resolve().parents[2]
_FORMULA_ENGINE_SRC = _REPO_ROOT / "src" / "product_a" / "phase_a4" / "formula_evaluator.py"

# _exrs_range_utils.py NÃO é uma cópia vendorizada de normalizer.py — normalizer.py tem
# efeitos colaterais de import (sys.path.insert, import de pipeline_contracts) que quebram
# em qualquer máquina sem o repositório EXRS completo, o que anula a proposta de "standalone"
# do .py gerado. Este é um template estático e autocontido com APENAS expand_range e suas
# dependências transitivas (parse_cell_coordinate, FULL_COLUMN_PATTERN) — sem sys.path,
# sem pipeline_contracts.
_RANGE_UTILS_SRC_TEXT = '''"""
Módulo vendorizado standalone — expand_range (extraído de src/kernel/phase_a1_5/normalizer.py).
Sem dependências do repositório EXRS: só `re` e `openpyxl.utils`.
"""
import re

from openpyxl.utils import get_column_letter, column_index_from_string

FULL_COLUMN_PATTERN = re.compile(r'^([A-Z]+):([A-Z]+)$', re.IGNORECASE)


def parse_cell_coordinate(coord: str) -> tuple[str, int]:
    """Separa 'A1' em ('A', 1)."""
    match = re.match(r'^([A-Z]+)(\\d+)$', coord.upper().replace('$', ''))
    if not match:
        raise ValueError(f"Coordenada inválida: {coord}")
    return match.group(1), int(match.group(2))


def expand_range(range_str: str, max_row: int | None = None) -> list[str]:
    """Expande um range como 'A1:B2' em lista de coordenadas individuais."""
    range_str = range_str.replace('$', '')

    full_col_match = FULL_COLUMN_PATTERN.match(range_str)
    if full_col_match:
        if max_row is None:
            max_row = 2000
        col1 = column_index_from_string(full_col_match.group(1))
        col2 = column_index_from_string(full_col_match.group(2))
        start_col = min(col1, col2)
        end_col = max(col1, col2)
        cells: list[str] = []
        for row in range(1, max_row + 1):
            for col in range(start_col, end_col + 1):
                cells.append(f"{get_column_letter(col)}{row}")
        return cells

    cells: list[str] = []
    if ':' in range_str:
        parts = range_str.split(':')
        if len(parts) != 2:
            return cells
        try:
            col1, row1 = parse_cell_coordinate(parts[0])
            col2, row2 = parse_cell_coordinate(parts[1])
        except ValueError:
            return cells

        col1_idx = column_index_from_string(col1)
        col2_idx = column_index_from_string(col2)
        start_col = min(col1_idx, col2_idx)
        end_col = max(col1_idx, col2_idx)
        start_row = min(row1, row2)
        end_row = max(row1, row2)

        if end_row - start_row > 2000:
            end_row = start_row + 2000

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cells.append(f"{get_column_letter(col)}{row}")
    else:
        cells.append(range_str)

    return cells
'''


def write_clean_output(
    job_output_dir: Path,
    stem: str,
    certified,
    dag,
    fmap,
    norm_ir,
    dest_dir: Path,
    debug: bool = False,
) -> Path:
    """Monta dest_dir com o .py replay + engine vendorizado + relatório .html.
    Se debug=True, também copia os JSONs técnicos de job_output_dir para dest_dir."""
    dest_dir.mkdir(parents=True, exist_ok=True)

    source = render_replay_module(dag, fmap, norm_ir, source_file=f"{stem}.xlsx")
    (dest_dir / f"{stem}.py").write_text(source, encoding="utf-8")

    if not _FORMULA_ENGINE_SRC.exists():
        raise RuntimeError(
            f"EXRS vendored engine source não encontrado: {_FORMULA_ENGINE_SRC}. "
            "Repositório pode estar com layout alterado ou instalação corrompida."
        )
    shutil.copy2(_FORMULA_ENGINE_SRC, dest_dir / "_exrs_formula_engine.py")
    (dest_dir / "_exrs_range_utils.py").write_text(_RANGE_UTILS_SRC_TEXT, encoding="utf-8")

    generate_html_report(
        certified.validation_report.results,
        dest_dir / f"{stem}_report.html",
        fmap=fmap,
        source_file=f"{stem}.xlsx",
        title=f"EXRS — Relatório de {stem}",
    )

    if debug:
        for json_file in job_output_dir.glob(f"{stem}_*.json"):
            shutil.copy2(json_file, dest_dir / json_file.name)

    return dest_dir
